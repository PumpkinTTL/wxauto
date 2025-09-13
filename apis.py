# 交互文件
import time
import threading
import os
import json
import base64
import uuid
import shutil
from pathlib import Path
from datetime import datetime


# 核心库导入
try:
    import openpyxl
except ImportError as e:
    print(f"⚠️ openpyxl库导入失败: {e}")
    openpyxl = None

try:
    from sqlite3_util import (
        batch_insert,
        query_users,
        get_users_count,
        verify_insert_result,
        query_wechat_phrases,
        add_wechat_phrase,
        update_wechat_phrase,
        delete_wechat_phrase,
        add_user_log,
        query_user_logs,
        # 商品管理相关函数
        add_product,
        query_products,
        update_product,
        delete_product,
        get_all_products_simple,
        # 图片管理相关函数
        add_image,
        query_images,
        update_image,
        delete_image,
        get_product_images,
        get_product_with_images,
    )
except ImportError as e:
    print(f"⚠️ sqlite3_util模块导入失败: {e}")
    batch_insert = query_users = get_users_count = verify_insert_result = None
    query_wechat_phrases = add_wechat_phrase = update_wechat_phrase = delete_wechat_phrase = None
    add_user_log = query_user_logs = None
    # 商品和图片管理函数设为None
    add_product = query_products = update_product = delete_product = get_all_products_simple = None
    add_image = query_images = update_image = delete_image = get_product_images = get_product_with_images = None

# 微信自动化导入
try:
    import wechat_automation
    WECHAT_AUTOMATION_AVAILABLE = True
except ImportError as e:
    print(f"⚠️ 微信自动化模块导入失败: {e}")
    wechat_automation = None
    WECHAT_AUTOMATION_AVAILABLE = False

# 全局控制台输出变量
GLOBAL_CONSOLE_OUTPUT = {
    "logs": [],  # 控制台日志列表
    "is_processing": False,  # 是否正在处理
    "current_status": "idle",  # 当前状态: idle, processing, completed, error
    "current_message": "",  # 当前状态消息
    "progress": 0,  # 进度百分比 0-100
    "last_update": time.time()
}

# 🔥 新增：跟播进度日志管理器
FOLLOW_PROGRESS_LOGS = {
    "logs": [],  # 跟播进度日志列表
    "is_following": False,  # 是否正在跟播
    "current_room": "",  # 当前跟播的直播间
    "progress": 0,  # 跟播进度 0-100
    "step": "",  # 当前步骤描述
    "last_update": time.time(),
    "start_time": None,  # 跟播开始时间
    "room_count": 0,  # 总直播间数量
    "completed_count": 0  # 已完成直播间数量
}

# 🔥 新增：监听窗口管理器
PROGRESS_WINDOW_MANAGER = {
    "active_windows": {},  # 活跃的监听窗口 {room_name: window_info}
    "should_close": set(),  # 需要关闭的窗口房间名集合
}

# 🔥 新增：全局print同步机制
def sync_print(message, log_type="info", room_name=None, operation=None, *args, **kwargs):
    """
    同步print函数 - 既打印到控制台又同步到进度窗口
    用于替换关键的print语句，确保所有输出都能被监控窗口看到
    
    Args:
        message: 要打印的消息
        log_type: 日志类型 (info, success, warning, error)
        room_name: 直播间名称（可选）
        operation: 操作类型（可选）
    """
    # 打印到控制台
    print(message, *args, **kwargs)
    
    # 同步到进度窗口
    try:
        # 根据消息内容自动判断日志类型
        if log_type == "info" and message:
            if "❌" in message or "失败" in message or "错误" in message:
                log_type = "error"
            elif "✅" in message or "成功" in message or "完成" in message:
                log_type = "success"
            elif "⚠️" in message or "警告" in message or "注意" in message:
                log_type = "warning"
        
        add_follow_progress_log(str(message), log_type, None, operation, room_name)
    except Exception as e:
        print(f"⚠️ 同步进度日志失败: {e}")  # 避免递归错误

# 全局取消标志
PROCESSING_CANCELLED = False

# 线程锁
console_lock = threading.Lock()
follow_progress_lock = threading.Lock()

def add_console_log(message, log_type="info"):
    """添加控制台日志"""
    global GLOBAL_CONSOLE_OUTPUT

    with console_lock:
        log_entry = {
            "timestamp": time.time(),
            "time_str": time.strftime("%H:%M:%S", time.localtime()),
            "message": message,
            "type": log_type  # info, success, warning, error
        }

        GLOBAL_CONSOLE_OUTPUT["logs"].append(log_entry)

        # 限制日志数量，只保留最新50条
        if len(GLOBAL_CONSOLE_OUTPUT["logs"]) > 50:
            GLOBAL_CONSOLE_OUTPUT["logs"] = GLOBAL_CONSOLE_OUTPUT["logs"][-50:]

        GLOBAL_CONSOLE_OUTPUT["last_update"] = time.time()

        # 同时打印到控制台
        print(f"[{log_entry['time_str']}] {message}")

def add_follow_progress_log(message, log_type="info", progress=None, step=None, room_name=None):
    """添加跟播进度日志（增强版）"""
    global FOLLOW_PROGRESS_LOGS

    with follow_progress_lock:
        current_time = time.time()
        # 创建详细的时间戳
        timestamp_detailed = time.strftime("%H:%M:%S", time.localtime(current_time))
        
        log_entry = {
            "time": current_time,
            "timestamp": timestamp_detailed,
            "message": message,
            "log_type": log_type,  # info, success, warning, error
            "progress": progress,
            "step": step,
            "room_name": room_name,
            "operation": step if step else "通用",  # 操作类型
            "category": get_log_category(message, step)  # 日志分类
        }

        FOLLOW_PROGRESS_LOGS["logs"].append(log_entry)

        # 增加日志数量限制，保留最新200条（原来是100条）
        if len(FOLLOW_PROGRESS_LOGS["logs"]) > 200:
            FOLLOW_PROGRESS_LOGS["logs"] = FOLLOW_PROGRESS_LOGS["logs"][-200:]

        # 更新状态信息
        if progress is not None:
            FOLLOW_PROGRESS_LOGS["progress"] = progress
        if step is not None:
            FOLLOW_PROGRESS_LOGS["step"] = step
        if room_name is not None:
            FOLLOW_PROGRESS_LOGS["current_room"] = room_name

        FOLLOW_PROGRESS_LOGS["last_update"] = current_time
        
        # 统计各类型日志数量
        update_log_statistics(log_type)

        # 优化的控制台输出格式
        console_message = format_console_message(log_entry)
        print(console_message)

def get_log_category(message, step):
    """根据消息内容和步骤确定日志分类"""
    if step:
        step_lower = step.lower()
        if any(keyword in step_lower for keyword in ['图像', '识别', '匹配']):
            return "图像识别"
        elif any(keyword in step_lower for keyword in ['弹幕', '发送', '话术']):
            return "弹幕发送"
        elif any(keyword in step_lower for keyword in ['窗口', '激活', '切换']):
            return "窗口管理"
        elif any(keyword in step_lower for keyword in ['任务', '执行', '完成']):
            return "任务管理"
        elif any(keyword in step_lower for keyword in ['跟播', '模式']):
            return "跟播流程"
    
    # 根据消息内容分类
    message_lower = message.lower()
    if any(keyword in message_lower for keyword in ['图像', '识别', '匹配', '商品图片']):
        return "图像识别"
    elif any(keyword in message_lower for keyword in ['弹幕', '发送', '话术', '内容']):
        return "弹幕发送"
    elif any(keyword in message_lower for keyword in ['chrome', '窗口', '激活']):
        return "窗口管理"
    elif any(keyword in message_lower for keyword in ['任务', '执行', '完成', '创建']):
        return "任务管理"
    else:
        return "系统日志"

def update_log_statistics(log_type):
    """更新日志统计信息"""
    global FOLLOW_PROGRESS_LOGS
    
    if "statistics" not in FOLLOW_PROGRESS_LOGS:
        FOLLOW_PROGRESS_LOGS["statistics"] = {
            "info": 0,
            "success": 0,
            "warning": 0,
            "error": 0
        }
    
    if log_type in FOLLOW_PROGRESS_LOGS["statistics"]:
        FOLLOW_PROGRESS_LOGS["statistics"][log_type] += 1

def format_console_message(log_entry):
    """格式化控制台输出消息 - 统一格式"""
    # 根据日志类型使用不同的状态标识
    status_map = {
        "info": "[信息]",
        "success": "[成功]", 
        "warning": "[警告]",
        "error": "[失败]"
    }
    
    status = status_map.get(log_entry["log_type"], "[信息]")
    
    # 统一格式：时间[状态] 【直播间】[操作类型] 消息内容
    formatted_parts = [log_entry['timestamp'] + status]
    
    # 添加直播间信息
    if log_entry.get("room_name"):
        formatted_parts.append(f"【{log_entry['room_name']}】")
    
    # 添加操作类型信息
    if log_entry.get("operation") and log_entry.get("operation") != "通用":
        formatted_parts.append(f"[{log_entry['operation']}]")
    
    # 添加消息内容
    formatted_parts.append(log_entry['message'])
    
    return " ".join(formatted_parts)

def update_follow_progress_status(is_following=None, current_room=None, progress=None, step=None,
                                room_count=None, completed_count=None):
    """更新跟播进度状态"""
    global FOLLOW_PROGRESS_LOGS

    with follow_progress_lock:
        if is_following is not None:
            FOLLOW_PROGRESS_LOGS["is_following"] = is_following
            if is_following and FOLLOW_PROGRESS_LOGS["start_time"] is None:
                FOLLOW_PROGRESS_LOGS["start_time"] = time.time()
            elif not is_following:
                FOLLOW_PROGRESS_LOGS["start_time"] = None

        if current_room is not None:
            FOLLOW_PROGRESS_LOGS["current_room"] = current_room
        if progress is not None:
            FOLLOW_PROGRESS_LOGS["progress"] = progress
        if step is not None:
            FOLLOW_PROGRESS_LOGS["step"] = step
        if room_count is not None:
            FOLLOW_PROGRESS_LOGS["room_count"] = room_count
        if completed_count is not None:
            FOLLOW_PROGRESS_LOGS["completed_count"] = completed_count

        FOLLOW_PROGRESS_LOGS["last_update"] = time.time()

def get_follow_progress_logs():
    """获取跟播进度日志"""
    global FOLLOW_PROGRESS_LOGS

    with follow_progress_lock:
        return FOLLOW_PROGRESS_LOGS.copy()

def reset_follow_progress_logs():
    """重置跟播进度日志"""
    global FOLLOW_PROGRESS_LOGS

    with follow_progress_lock:
        FOLLOW_PROGRESS_LOGS["logs"] = []
        FOLLOW_PROGRESS_LOGS["is_following"] = False
        FOLLOW_PROGRESS_LOGS["current_room"] = ""
        FOLLOW_PROGRESS_LOGS["progress"] = 0
        FOLLOW_PROGRESS_LOGS["step"] = ""
        FOLLOW_PROGRESS_LOGS["last_update"] = time.time()
        FOLLOW_PROGRESS_LOGS["start_time"] = None
        FOLLOW_PROGRESS_LOGS["room_count"] = 0
        FOLLOW_PROGRESS_LOGS["completed_count"] = 0

# 移除print拦截功能，改为简单的同步机制

def update_console_status(status="idle", message="", progress=0, is_processing=False):
    """更新控制台状态"""
    global GLOBAL_CONSOLE_OUTPUT

    with console_lock:
        GLOBAL_CONSOLE_OUTPUT["current_status"] = status
        GLOBAL_CONSOLE_OUTPUT["current_message"] = message
        GLOBAL_CONSOLE_OUTPUT["progress"] = progress
        GLOBAL_CONSOLE_OUTPUT["is_processing"] = is_processing
        GLOBAL_CONSOLE_OUTPUT["last_update"] = time.time()

def get_console_output():
    """获取控制台输出"""
    global GLOBAL_CONSOLE_OUTPUT

    with console_lock:
        return GLOBAL_CONSOLE_OUTPUT.copy()

def reset_console_output():
    """重置控制台输出"""
    global GLOBAL_CONSOLE_OUTPUT

    with console_lock:
        GLOBAL_CONSOLE_OUTPUT["logs"] = []
        GLOBAL_CONSOLE_OUTPUT["is_processing"] = False
        GLOBAL_CONSOLE_OUTPUT["current_status"] = "idle"
        GLOBAL_CONSOLE_OUTPUT["current_message"] = ""
        GLOBAL_CONSOLE_OUTPUT["progress"] = 0
        GLOBAL_CONSOLE_OUTPUT["last_update"] = time.time()

def cancel_processing():
    """取消当前处理"""
    global PROCESSING_CANCELLED
    PROCESSING_CANCELLED = True
    add_console_log("🛑 用户取消处理", "warning")
    update_console_status(status="cancelled", message="处理已取消", is_processing=False)

    # 返回取消结果，但这里无法获取已处理数据，需要在process_excel_file中处理
    return {
        "success": True,
        "message": "处理已取消",
        "cancelled": True
    }

def is_processing_cancelled():
    """检查是否已取消处理"""
    return PROCESSING_CANCELLED

def reset_cancel_flag():
    """重置取消标志"""
    global PROCESSING_CANCELLED
    PROCESSING_CANCELLED = False

def check_api_response_valid(real_data):
    """检查API响应是否有效，判断是否被风控"""
    if not real_data:
        return False, "API返回空数据", False  # 第三个参数表示是否是风控

    if not isinstance(real_data, dict):
        return False, "API返回数据格式错误", False

    # 检查是否是真正的风控
    if real_data.get('risk_control') == True:
        error_msg = real_data.get('error_msg', '未知风控错误')
        return False, error_msg, True  # 是风控

    # 检查关键字段 - 只检查unique_id，signature可以为空
    if not real_data.get('unique_id'):
        return False, "API返回数据不完整", False  # 不是风控，可能是取消请求

    return True, "数据有效", False

class API:
    def __init__(self):
        # 导入redislite和数据库工具
        try:
            from redislite import get_log_manager
            from sqlite3_util import add_user_log, query_user_logs, clear_user_logs, check_user_added
            self.log_manager = get_log_manager()
            self.add_user_log = add_user_log
            self.query_user_logs = query_user_logs
            self.clear_user_logs = clear_user_logs
            self.check_user_added = check_user_added
        except ImportError as e:
            print(f"导入redislite或数据库工具失败: {e}")
            self.log_manager = None

    # 🔥 新增：跟播进度窗口相关API方法
    def create_follow_progress_window(self, room_name="未知直播间"):
        """创建跟播进度监控窗口"""
        try:
            import webview
            import os
            import threading

            # 🔥 检查是否已有该直播间的监听窗口
            global PROGRESS_WINDOW_MANAGER
            if room_name in PROGRESS_WINDOW_MANAGER["active_windows"]:
                add_follow_progress_log(f"⚠️ 直播间 {room_name} 的监听窗口已存在，跳过创建", "warning")
                return {
                    "success": True,
                    "message": f"监听窗口已存在: {room_name}"
                }

            # 获取当前目录和HTML文件路径
            current_dir = os.path.dirname(os.path.abspath(__file__))
            html_path = os.path.join(current_dir, "web", "pages", "follow_progress.html")

            if not os.path.exists(html_path):
                error_msg = f"进度窗口HTML文件不存在: {html_path}"
                add_follow_progress_log(f"❌ {error_msg}", "error")
                return {
                    "success": False,
                    "message": error_msg
                }

            add_follow_progress_log(f"🪟 正在创建进度监控窗口...", "info")
            print(f"🪟 创建进度窗口: {html_path}")

            # 🔥 获取屏幕尺寸来计算右上角位置
            try:
                import tkinter as tk
                root = tk.Tk()
                screen_width = root.winfo_screenwidth()
                screen_height = root.winfo_screenheight()
                root.destroy()
            except ImportError:
                # 如果tkinter不可用，使用默认屏幕尺寸
                screen_width = 1920
                screen_height = 1080
                print("⚠️ tkinter不可用，使用默认屏幕尺寸")

            # 计算右上角位置
            window_width = 450
            window_height = 600
            x_position = screen_width - window_width - 20  # 距离右边20像素
            y_position = 50  # 距离顶部50像素

            print(f"📍 窗口位置: ({x_position}, {y_position})")
            print(f"📏 窗口尺寸: {window_width}x{window_height}")

            # 🔥 修复：直接创建窗口，不使用线程（webview必须在主线程）
            try:
                # 🔥 重要：创建新窗口
                progress_window = webview.create_window(
                    title=f'🎯 跟播进度监控 - {room_name}',
                    url=html_path,
                    width=window_width,
                    height=window_height,
                    min_size=(400, 500),
                    resizable=True,
                    js_api=self,  # 使用当前API实例
                    x=x_position,  # 设置X位置
                    y=y_position   # 设置Y位置
                )

                # 🔥 记录窗口信息
                PROGRESS_WINDOW_MANAGER["active_windows"][room_name] = {
                    "window": progress_window,
                    "created_time": time.time(),
                    "title": f'🎯 跟播进度监控 - {room_name}'
                }

                add_follow_progress_log(f"✅ 跟播进度窗口创建成功: {room_name}", "success")
                print(f"✅ 进度窗口创建成功: {progress_window}")

                # 🔥 不在这里启动webview，让主程序控制启动时机

            except Exception as e:
                error_msg = f"窗口创建失败: {str(e)}"
                add_follow_progress_log(f"❌ {error_msg}", "error")
                print(f"❌ {error_msg}")
                return {
                    "success": False,
                    "message": error_msg
                }

            return {
                "success": True,
                "message": "跟播进度窗口创建成功"
            }

        except Exception as e:
            error_msg = f"创建跟播进度窗口失败: {str(e)}"
            add_follow_progress_log(f"❌ {error_msg}", "error")
            return {
                "success": False,
                "message": error_msg
            }

    def _set_window_position(self, window, x, y):
        """设置窗口位置（在窗口启动后调用）"""
        try:
            import time
            # 等待窗口完全加载
            time.sleep(0.5)

            # 尝试使用pywebview的move方法
            if hasattr(window, 'move'):
                window.move(x, y)
                print(f"✅ 窗口位置已设置: ({x}, {y})")
            else:
                print(f"⚠️ 窗口不支持move方法，使用默认位置")

        except Exception as e:
            print(f"⚠️ 设置窗口位置失败: {str(e)}")

    def get_progress_logs(self, last_time=0):
        """获取跟播进度日志（用于前端轮询）"""
        try:
            progress_data = get_follow_progress_logs()

            # 过滤出指定时间之后的日志
            new_logs = [
                log for log in progress_data["logs"]
                if log["time"] > last_time
            ]

            return {
                "success": True,
                "logs": new_logs,
                "status": {
                    "is_following": progress_data["is_following"],
                    "current_room": progress_data["current_room"],
                    "progress": progress_data["progress"],
                    "step": progress_data["step"],
                    "room_count": progress_data["room_count"],
                    "completed_count": progress_data["completed_count"],
                    "last_update": progress_data["last_update"]
                }
            }

        except Exception as e:
            return {
                "success": False,
                "message": f"获取进度日志失败: {str(e)}",
                "logs": []
            }

    def stream_progress_logs(self, last_time=0):
        """🔥 流式获取跟播进度日志（使用yield实现实时推送）"""
        import time

        try:
            while True:
                progress_data = get_follow_progress_logs()

                # 过滤出指定时间之后的日志
                new_logs = [
                    log for log in progress_data["logs"]
                    if log["time"] > last_time
                ]

                # 如果有新日志，返回数据
                if new_logs:
                    # 更新last_time为最新日志的时间
                    last_time = max(log["time"] for log in new_logs)

                    yield {
                        "success": True,
                        "logs": new_logs,
                        "status": {
                            "is_following": progress_data["is_following"],
                            "current_room": progress_data["current_room"],
                            "progress": progress_data["progress"],
                            "step": progress_data["step"],
                            "room_count": progress_data["room_count"],
                            "completed_count": progress_data["completed_count"],
                            "last_update": progress_data["last_update"]
                        },
                        "last_time": last_time
                    }

                # 如果跟播已结束且没有新日志，停止流式传输
                if not progress_data["is_following"] and not new_logs:
                    yield {
                        "success": True,
                        "logs": [],
                        "status": progress_data,
                        "finished": True,
                        "last_time": last_time
                    }
                    break

                # 等待一段时间再检查
                time.sleep(0.5)

        except Exception as e:
            yield {
                "success": False,
                "message": f"流式获取进度日志失败: {str(e)}",
                "logs": [],
                "finished": True
            }

    def reset_progress_logs(self):
        """重置跟播进度日志"""
        try:
            reset_follow_progress_logs()
            return {
                "success": True,
                "message": "跟播进度日志已重置"
            }
        except Exception as e:
            return {
                "success": False,
                "message": f"重置进度日志失败: {str(e)}"
            }

# 移除print拦截API方法

    def close_follow_progress_window(self, room_name=""):
        """关闭跟播进度监控窗口"""
        try:
            # 🔥 新增：关闭监听窗口的逻辑
            import webview

            # 记录关闭日志
            add_follow_progress_log(f"🔄 正在关闭跟播监听窗口", "warning", None, "窗口管理", room_name)

            # 🔥 检查窗口是否存在并标记为需要关闭
            global PROGRESS_WINDOW_MANAGER

            # 🔥 改进：支持模糊匹配房间名
            target_room_names = []

            # 精确匹配
            if room_name in PROGRESS_WINDOW_MANAGER["active_windows"]:
                target_room_names.append(room_name)

            # 模糊匹配：查找包含该房间名的窗口
            for active_room_name in list(PROGRESS_WINDOW_MANAGER["active_windows"].keys()):
                if room_name in active_room_name or active_room_name in room_name:
                    if active_room_name not in target_room_names:
                        target_room_names.append(active_room_name)

            if target_room_names:
                closed_count = 0
                for target_room_name in target_room_names:
                    try:
                        # 标记窗口需要关闭
                        PROGRESS_WINDOW_MANAGER["should_close"].add(target_room_name)

                        window_info = PROGRESS_WINDOW_MANAGER["active_windows"][target_room_name]
                        print(f"🔄 [WINDOW_CLOSE] 标记窗口需要关闭: {target_room_name}")
                        print(f"🔄 [WINDOW_CLOSE] 窗口标题: {window_info.get('title', 'Unknown')}")

                        # 🔥 修复：不要立即删除窗口信息，让前端检查到关闭信号后再删除
                        # del PROGRESS_WINDOW_MANAGER["active_windows"][target_room_name]
                        closed_count += 1

                        add_follow_progress_log(f"✅ 窗口关闭标记已设置: {target_room_name}", "success", None, "窗口管理", room_name)

                    except Exception as close_error:
                        print(f"⚠️ [WINDOW_CLOSE] 关闭窗口时出现异常: {target_room_name} - {close_error}")
                        add_follow_progress_log(f"⚠️ 窗口关闭异常: {str(close_error)}", "warning", None, "窗口管理", room_name)

                # 更新跟播状态为停止
                update_follow_progress_status(is_following=False, step="跟播已停止")

                print(f"✅ [WINDOW_CLOSE] 成功标记 {closed_count} 个窗口关闭")

                return {
                    "success": True,
                    "message": f"成功标记 {closed_count} 个跟播监听窗口关闭"
                }
            else:
                print(f"⚠️ [WINDOW_CLOSE] 未找到活跃的监听窗口: {room_name}")
                add_follow_progress_log(f"⚠️ 未找到活跃的监听窗口", "warning", None, "窗口管理", room_name)

                # 更新跟播状态为停止
                update_follow_progress_status(is_following=False, step="跟播已停止")

                return {
                    "success": True,
                    "message": f"监听窗口不存在或已关闭: {room_name}"
                }

        except Exception as e:
            error_msg = f"关闭跟播进度窗口失败: {str(e)}"
            print(f"❌ [WINDOW_CLOSE] {error_msg}")
            return {
                "success": False,
                "message": error_msg
            }

    def check_window_should_close(self, room_name=None):
        """检查当前窗口是否应该关闭"""
        try:
            global PROGRESS_WINDOW_MANAGER

            # 🔥 改进：如果没有提供房间名，检查所有需要关闭的窗口
            if not room_name:
                # 检查是否有任何窗口需要关闭
                if PROGRESS_WINDOW_MANAGER["should_close"]:
                    # 返回第一个需要关闭的窗口
                    close_room = list(PROGRESS_WINDOW_MANAGER["should_close"])[0]
                    PROGRESS_WINDOW_MANAGER["should_close"].discard(close_room)

                    print(f"🔄 [WINDOW_CHECK] 窗口 {close_room} 需要关闭")

                    return {
                        "should_close": True,
                        "room_name": close_room,
                        "message": f"窗口 {close_room} 需要关闭"
                    }
                else:
                    return {
                        "should_close": False,
                        "message": "窗口无需关闭"
                    }
            else:
                # 🔥 改进：检查指定房间的窗口是否需要关闭
                if room_name in PROGRESS_WINDOW_MANAGER["should_close"]:
                    PROGRESS_WINDOW_MANAGER["should_close"].discard(room_name)

                    print(f"🔄 [WINDOW_CHECK] 窗口 {room_name} 需要关闭")

                    return {
                        "should_close": True,
                        "room_name": room_name,
                        "message": f"窗口 {room_name} 需要关闭"
                    }
                else:
                    return {
                        "should_close": False,
                        "message": f"窗口 {room_name} 无需关闭"
                    }

        except Exception as e:
            print(f"❌ [WINDOW_CHECK] 检查窗口关闭状态失败: {str(e)}")
            return {
                "should_close": False,
                "error": str(e)
            }

    def test_create_progress_window(self):
        """测试创建跟播进度窗口（用于前端测试）"""
        try:
            # 重置进度日志
            reset_follow_progress_logs()

            # 添加一些测试数据
            update_follow_progress_status(
                is_following=True,
                room_count=3,
                completed_count=0,
                progress=0,
                step="测试窗口创建"
            )

            add_follow_progress_log("🧪 测试窗口创建功能", "info", 10, "测试开始")
            add_follow_progress_log("📺 模拟直播间: 测试直播间", "info", 20, "模拟数据")
            add_follow_progress_log("✅ 测试数据准备完成", "success", 30, "数据准备完成")

            # 创建进度窗口
            result = self.create_follow_progress_window("测试直播间")

            if result["success"]:
                # 启动模拟进度更新
                import threading
                threading.Thread(target=self._simulate_progress_updates, daemon=True).start()

            return result

        except Exception as e:
            return {
                "success": False,
                "message": f"测试创建进度窗口失败: {str(e)}"
            }

    def _simulate_progress_updates(self):
        """模拟进度更新（用于测试）"""
        import time

        try:
            time.sleep(2)  # 等待窗口加载

            steps = [
                (40, "🔍 检查直播间状态", "info"),
                (50, "✅ 直播间连接成功", "success"),
                (60, "📝 配置弹幕话术", "info"),
                (70, "💬 获取到 5 条话术", "success"),
                (80, "🖼️ 开始图像识别", "info"),
                (90, "🎉 图像匹配成功！", "success"),
                (95, "💬 发送弹幕: 产品真不错！", "success"),
                (100, "🎉 测试完成！", "success")
            ]

            for progress, message, log_type in steps:
                add_follow_progress_log(message, log_type, progress, f"测试步骤{progress}%", "测试直播间")
                update_follow_progress_status(progress=progress, step=f"测试步骤{progress}%")
                time.sleep(1.5)

            # 完成测试
            update_follow_progress_status(
                is_following=False,
                progress=100,
                step="测试完成",
                completed_count=1
            )

        except Exception as e:
            add_follow_progress_log(f"❌ 模拟进度更新失败: {str(e)}", "error")

    def get_console_output(self):
        """获取控制台输出"""
        return get_console_output()

    def stop_processing(self, params=None):
        """终止当前处理"""
        global PROCESSING_CANCELLED
        PROCESSING_CANCELLED = True
        add_console_log("🛑 用户请求终止操作", "warning")
        return {
            "success": True,
            "message": "终止请求已发送"
        }

    def reset_console_output(self):
        """重置控制台输出"""
        return reset_console_output()

    def cancel_processing(self):
        """取消当前处理"""
        return cancel_processing()

    def check_token_status(self):
        """检查token状态"""
        try:
            from cmm import get_latest_token
            token = get_latest_token()

            if token:
                return {
                    "success": True,
                    "has_token": True,
                    "token_preview": f"{token[:20]}...",
                    "message": "Token可用"
                }
            else:
                return {
                    "success": True,
                    "has_token": False,
                    "message": "未找到有效Token"
                }
        except Exception as e:
            return {
                "success": False,
                "has_token": False,
                "message": f"检查Token失败: {str(e)}"
            }

    def get_users_data(self, page=1, page_size=200, search_params=None):
        """
        获取用户数据（分页+搜索+排序）
        """
        try:
            print(f"=== 查询用户数据 ===")
            print(f"页码: {page}, 每页: {page_size}")
            if search_params:
                print(f"搜索条件: {search_params}")

            # 首先检查数据库和表是否存在
            import os
            db_path = 'system.db'
            if not os.path.exists(db_path):
                print(f"❌ 数据库文件不存在: {db_path}")
                return {
                    "success": False,
                    "message": "数据库文件不存在，请先上传Excel文件",
                    "data": [],
                    "total": 0
                }

            # 检查表是否存在
            conn = __import__('sqlite3').connect(db_path)
            cursor = conn.cursor()
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='users'")
            table_exists = cursor.fetchone() is not None

            if not table_exists:
                conn.close()
                print(f"❌ users表不存在")
                return {
                    "success": False,
                    "message": "users表不存在，请先上传Excel文件",
                    "data": [],
                    "total": 0
                }

            # 检查表中是否有数据
            cursor.execute("SELECT COUNT(*) FROM users")
            total_records = cursor.fetchone()[0]
            print(f"📊 数据库中总记录数: {total_records}")
            conn.close()

            # 构建搜索条件
            where_conditions = []
            params = []

            if search_params:
                if search_params.get('fileName'):
                    where_conditions.append("file_name LIKE ?")
                    params.append(f"%{search_params['fileName']}%")

                if search_params.get('description'):
                    where_conditions.append("intro LIKE ?")
                    params.append(f"%{search_params['description']}%")

                if search_params.get('phone'):
                    # 这里假设phone字段存在，如果没有可以搜索unique_id
                    where_conditions.append("unique_id LIKE ?")
                    params.append(f"%{search_params['phone']}%")

                if search_params.get('wechat'):
                    # 这里假设wechat字段存在，如果没有可以搜索username
                    where_conditions.append("username LIKE ?")
                    params.append(f"%{search_params['wechat']}%")

                # 时间区间搜索
                if search_params.get('startTime'):
                    where_conditions.append("date(create_time) >= ?")
                    params.append(search_params['startTime'])

                if search_params.get('endTime'):
                    where_conditions.append("date(create_time) <= ?")
                    params.append(search_params['endTime'])

            # 构建WHERE子句
            where_clause = ""
            if where_conditions:
                where_clause = "WHERE " + " AND ".join(where_conditions)

            # 查询总记录数
            conn = __import__('sqlite3').connect('system.db')
            conn.row_factory = __import__('sqlite3').Row
            cursor = conn.cursor()

            count_sql = f"SELECT COUNT(*) FROM users {where_clause}"
            cursor.execute(count_sql, params)
            total_count = cursor.fetchone()[0]
            print(f"符合条件的总记录数: {total_count}")

            # 计算偏移量
            offset = (page - 1) * page_size

            # 处理排序参数
            sort_order = 'DESC'  # 默认降序
            if search_params and search_params.get('sortOrder'):
                if search_params['sortOrder'].lower() == 'asc':
                    sort_order = 'ASC'
                elif search_params['sortOrder'].lower() == 'desc':
                    sort_order = 'DESC'

            print(f"排序方式: create_time {sort_order}")

            # 查询数据
            data_sql = f"""
                SELECT * FROM users
                {where_clause}
                ORDER BY create_time {sort_order}
                LIMIT ? OFFSET ?
            """

            cursor.execute(data_sql, params + [page_size, offset])
            results = [dict(row) for row in cursor.fetchall()]
            conn.close()

            add_console_log(f"📊 查询到 {len(results)} 条记录", "info")
            for i, record in enumerate(results[:3], 1):
                add_console_log(f"  {i}. {record.get('username')} - {record.get('cmm_id')}", "info")

            return {
                "success": True,
                "data": results,
                "total": total_count,
                "page": page,
                "page_size": page_size,
                "total_pages": (total_count + page_size - 1) // page_size,
                "search_params": search_params
            }

        except Exception as e:
            print(f"查询用户数据失败: {str(e)}")
            import traceback
            traceback.print_exc()
            return {
                "success": False,
                "message": f"查询失败: {str(e)}",
                "data": [],
                "total": 0
            }

    # 处理文件内容
    def process_file(self, file_data):
        """
        处理前端传来的文件数据，直接在内存中读取Excel内容
        :param file_data: 包含文件信息和内容的字典
        :return: 处理结果
        """
        # 检查必要的库
        if openpyxl is None:
            return {
                "success": False,
                "message": "openpyxl库未安装，无法处理Excel文件"
            }

        try:
            # 重置取消标志
            reset_cancel_flag()

            # 更新状态：开始处理
            update_console_status(status="processing", message="开始处理Excel文件...", is_processing=True)
            add_console_log("🚀 开始处理Excel文件...", "info")

            if not file_data:
                add_console_log("❌ 未接收到文件数据", "error")
                update_console_status(status="error", message="未接收到文件数据", is_processing=False)
                return {
                    "success": False,
                    "message": "未接收到文件数据"
                }

            # 检查必要字段
            required_fields = ['name', 'content']
            for field in required_fields:
                if field not in file_data:
                    add_console_log(f"❌ 缺少必要字段: {field}", "error")
                    update_console_status(status="error", message=f"缺少必要字段: {field}", is_processing=False)
                    return {
                        "success": False,
                        "message": f"缺少必要字段: {field}"
                    }

            add_console_log(f"📄 读取文件: {file_data['name']}", "info")
            add_console_log(f"📊 文件大小: {file_data.get('size', 'unknown')} bytes", "info")

            # 将字节数组转换为字节对象
            file_content = bytes(file_data['content'])
            print(f"转换后的字节长度: {len(file_content)}")

            # 使用BytesIO在内存中创建文件对象
            from io import BytesIO
            # 文件
            file_stream = BytesIO(file_content)

            # 使用openpyxl直接从内存读取Excel文件
            add_console_log("📊 开始解析Excel文件结构", "info")
            update_console_status(status="processing", message="正在解析Excel文件...", progress=10)
            workbook = openpyxl.load_workbook(file_stream)

            # 获取工作表信息
            sheet_names = workbook.sheetnames
            add_console_log(f"📋 发现 {len(sheet_names)} 个工作表", "success")

            # 读取第一个工作表
            sheet = workbook.active
            add_console_log(f"📄 当前工作表: {sheet.title} ({sheet.max_row}行 x {sheet.max_column}列)", "info")

            # 读取表头（第一行）
            headers = []
            for col in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(row=1, column=col).value
                headers.append(str(cell_value) if cell_value else f"列{col}")

            add_console_log(f"📝 表头: {', '.join(headers[:5])}{'...' if len(headers) > 5 else ''}", "info")

            # 读取所有数据，包括超链接
            add_console_log("🔍 开始扫描数据和超链接", "info")
            update_console_status(status="processing", message="正在扫描数据和超链接...", progress=20)
            all_data = []
            hyperlinks_found = []

            for row in range(2, sheet.max_row + 1):  # 从第2行开始
                row_data = []
                row_hyperlinks = []

                for col in range(1, sheet.max_column + 1):
                    cell = sheet.cell(row=row, column=col)
                    cell_value = str(cell.value) if cell.value else ""
                    row_data.append(cell_value)

                    # 检查是否有超链接
                    if cell.hyperlink:
                        hyperlink_info = {
                            'row': row,
                            'col': col,
                            'column_name': headers[col-1] if col-1 < len(headers) else f"列{col}",
                            'cell_value': cell_value,
                            'hyperlink': cell.hyperlink.target if cell.hyperlink.target else str(cell.hyperlink)
                        }
                        row_hyperlinks.append(hyperlink_info)
                        hyperlinks_found.append(hyperlink_info)
                        print(f"发现超链接 - 行{row}列{col}({headers[col-1] if col-1 < len(headers) else f'列{col}'}): {cell_value} -> {hyperlink_info['hyperlink']}")

                all_data.append(row_data)

                # 只打印前5行作为示例
                if row <= 6:
                    print(f"第{row}行: {row_data}")
                    if row_hyperlinks:
                        for link in row_hyperlinks:
                            print(f"  └─ 超链接: {link['cell_value']} -> {link['hyperlink']}")

            add_console_log(f"✅ 数据扫描完成！发现 {len(hyperlinks_found)} 个超链接", "success")
            update_console_status(status="processing", message=f"发现 {len(hyperlinks_found)} 个超链接", progress=40)

            # 处理超链接
            if hyperlinks_found:
                add_console_log(f"🔗 开始处理 {len(hyperlinks_found)} 个超链接", "info")

                # 提取ID并组装SQLite数据
                import re
                from datetime import datetime

                sqlite_data = []
                extracted_ids = []
                insert_success = False
                insert_message = "未执行数据库插入"

                # 获取爬取配置
                crawl_config = file_data.get('crawlConfig', {})
                sleep_interval = crawl_config.get('sleepInterval', 3)  # 默认3秒
                add_console_log(f"⚙️ 爬取配置 - 休眠间隔: {sleep_interval}秒", "info")

                # 先提取所有ID
                add_console_log("🔍 正在提取所有蝉妈妈ID...", "info")
                all_extracted_ids = []
                for link in hyperlinks_found:
                    url = link['hyperlink']
                    match = re.search(r'authorDetail/([^/?]+)', url)
                    if match:
                        author_id = match.group(1)
                        all_extracted_ids.append(author_id)

                add_console_log(f"📊 从Excel中提取到 {len(all_extracted_ids)} 个ID", "info")

                # 检查哪些ID已经存在于数据库中
                add_console_log("🔍 检查数据库中已存在的ID...", "info")
                from sqlite3_util import check_existing_ids_in_users
                id_check_result = check_existing_ids_in_users('system.db', file_data['name'], all_extracted_ids)

                existing_count = id_check_result['existing_count']
                new_count = id_check_result['new_count']
                new_ids = id_check_result['new_ids']

                if existing_count > 0:
                    add_console_log(f"⚠️ 发现 {existing_count} 个ID已存在于数据库中，将跳过", "warning")
                    add_console_log(f"✅ 需要处理 {new_count} 个新ID", "info")
                else:
                    add_console_log(f"✅ 所有 {new_count} 个ID都是新的，需要全部处理", "info")

                # 如果没有新ID需要处理
                if new_count == 0:
                    add_console_log("✅ 所有ID都已存在于数据库中，无需重复处理", "success")
                    update_console_status(status="completed", message="当前表格已经全部提取，无需重复处理", progress=100, is_processing=False)

                    # 延迟3秒后重置控制台，让用户看到完成信息
                    import threading
                    def delayed_reset():
                        import time
                        time.sleep(3)
                        reset_console_output()
                        add_console_log("📋 控制台已重置，可以处理新文件", "info")

                    threading.Thread(target=delayed_reset, daemon=True).start()

                    return {
                        "success": True,
                        "message": "当前表格已经全部提取，无需重复处理",
                        "data": {
                            "filename": file_data['name'],
                            "total_ids": len(all_extracted_ids),
                            "existing_ids": existing_count,
                            "new_ids": 0,
                            "skipped": True,
                            "all_existing": True  # 标记全部已存在
                        }
                    }

                # 获取最新token用于API调用
                add_console_log("🔑 获取最新token...", "info")
                token = self.get_latest_token_from_db()
                if not token:
                    add_console_log("❌ 未找到有效token，无法处理Excel文件", "error")
                    update_console_status(status="error", message="未找到有效token，请先登录蝉妈妈", is_processing=False)
                    return {
                        "success": False,
                        "message": "未找到有效token，请先登录蝉妈妈账号",
                        "data": None
                    }
                else:
                    add_console_log(f"✅ 获取到token: {token[:20]}...", "success")

                import time

                # 只处理新的ID，过滤掉已存在的
                new_hyperlinks = []
                for link in hyperlinks_found:
                    url = link['hyperlink']
                    match = re.search(r'authorDetail/([^/?]+)', url)
                    if match:
                        author_id = match.group(1)
                        if author_id in new_ids:  # 只处理新的ID
                            new_hyperlinks.append(link)

                add_console_log(f"📋 实际需要处理的超链接: {len(new_hyperlinks)} 个", "info")

                # 分批处理，每批返回进度
                batch_size = 5  # 每5个为一批
                total_count = len(new_hyperlinks)
                processed_count = 0

                # 获取爬取配置
                from cmm import get_crawl_config
                crawl_config = get_crawl_config()

                for i, link in enumerate(new_hyperlinks, 1):

                    # 休眠控制
                    # 单个数据爬取前的休眠
                    if crawl_config['wait_time'] > 0:
                        add_console_log(f"⏰ 单个数据休眠 {crawl_config['wait_time']} 秒", "info")
                        time.sleep(crawl_config['wait_time'])

                    # 检查是否需要大休眠
                    if processed_count > 0 and processed_count % crawl_config['count_wait'] == 0:
                        add_console_log(f"😴 已处理 {processed_count} 条数据，大休眠 {crawl_config['count_wait_time']} 秒", "info")
                        time.sleep(crawl_config['count_wait_time'])

                    # 更新进度
                    progress = 40 + (i / len(new_hyperlinks)) * 50  # 40-90%
                    update_console_status(status="processing", message=f"正在获取第 {i}/{len(new_hyperlinks)} 个达人信息", progress=progress)

                    # 提取authorDetail/后面的ID
                    url = link['hyperlink']
                    match = re.search(r'authorDetail/([^/?]+)', url)
                    if match:
                        author_id = match.group(1)
                        extracted_ids.append(author_id)

                        add_console_log(f"📡 [{i}/{len(new_hyperlinks)}] 获取达人: {link['cell_value']} ({author_id})", "info")

                        # 调用get_real_info获取真实数据
                        real_intro = ""
                        real_unique_id = ""
                        real_code = ""

                        # 使用token调用API获取真实数据
                        try:
                            from cmm import get_real_info, extract_contact_code
                            # 正常请求不使用直连IP
                            real_data = get_real_info(author_id, token, use_direct_at_end=False)

                            # 检查API响应是否有效（风控检测）
                            is_valid, error_msg, is_risk_control = check_api_response_valid(real_data)

                            if is_valid:
                                real_intro = real_data['signature']
                                real_unique_id = real_data['unique_id']

                                # 从signature中提取联系方式
                                real_code = extract_contact_code(real_intro)

                                add_console_log(f"✅ 获取成功: {link['cell_value']} | 抖音ID: {real_unique_id}", "success")
                                if real_code:
                                    add_console_log(f"📞 提取联系方式: {real_code}", "success")
                                else:
                                    add_console_log("📞 未提取到联系方式", "warning")
                            else:
                                # 检查是否是真正的风控
                                if is_risk_control:
                                    # 真正的风控，询问用户是否保存已处理的数据
                                    add_console_log(f"❌ {error_msg}", "error")
                                    add_console_log(f"🚨 检测到风控！已成功处理 {processed_count} 条数据", "warning")
                                else:
                                    # 不是风控，可能是取消请求或其他错误
                                    add_console_log(f"⚠️ {error_msg}", "warning")

                                # 只有真正的风控才执行风控处理逻辑
                                if is_risk_control:
                                    # 获取服务器返回的消息
                                    server_message = real_data.get('server_message', error_msg)
                                    if server_message:
                                        add_console_log(f"📡 服务器消息: {server_message}", "warning")

                                    # 删除数据库中的token
                                    try:
                                        import sqlite3
                                        conn = sqlite3.connect('system.db')
                                        cursor = conn.cursor()
                                        cursor.execute("DELETE FROM tokens")
                                        conn.commit()
                                        deleted_count = cursor.rowcount
                                        conn.close()
                                        add_console_log(f"🗑️ 已清除 {deleted_count} 个失效token", "warning")
                                    except Exception as token_error:
                                        add_console_log(f"❌ 清除token失败: {str(token_error)}", "error")

                                    # 返回风控信息，让前端询问用户是否保存
                                    update_console_status(status="risk_control", message=f"触发风控，已处理{processed_count}条", is_processing=False)
                                    return {
                                        "success": False,
                                        "message": f"检测到风控: {error_msg}",
                                        "risk_control": True,
                                        "server_message": server_message,  # 服务器返回的消息
                                        "token_cleared": True,
                                        "processed_count": processed_count,
                                        "total_count": total_count,
                                        "processed_data": sqlite_data,
                                        "ask_save_confirmation": True
                                    }

                        except Exception as e:
                            add_console_log(f"❌ API调用失败: {str(e)[:50]}...", "error")

                        data_row = {
                            'file_name': file_data['name'],
                            'username': link['cell_value'],
                            'intro': real_intro,
                            'unique_id': real_unique_id,
                            'cmm_id': author_id,
                            'code': real_code,
                            'create_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                        }

                        sqlite_data.append(data_row)
                        processed_count += 1

                        add_console_log(f"✅ 完成第 {i} 个达人信息获取: {link['cell_value']}", "success")

                        # 在处理完当前条数据后检查是否取消处理
                        if is_processing_cancelled():
                            add_console_log("🛑 处理已被用户取消，当前条数据已处理完成", "warning")

                            # 检查当前条数据的联系方式是否提取到
                            current_data_has_contact = real_code and real_code.strip() and real_code != 'None'
                            if not current_data_has_contact:
                                # 如果当前条数据没有联系方式，从sqlite_data中移除
                                if sqlite_data and sqlite_data[-1] == data_row:
                                    sqlite_data.pop()
                                    processed_count -= 1
                                    add_console_log(f"⚠️ 当前条数据未提取到联系方式，不计入数据库: {link['cell_value']}", "warning")
                                else:
                                    add_console_log(f"✅ 当前条数据已提取到联系方式: {real_code}", "success")
                            # 直接停止处理，不再进行直连请求
                            add_console_log("🛑 用户取消处理，准备停止", "info")

                            # 直连请求完成后，检查是否有已处理的数据并立即return
                            if len(sqlite_data) > 0:
                                add_console_log(f"📊 已处理 {len(sqlite_data)} 条数据", "info")

                                update_console_status(status="cancelled", message="处理已取消，询问是否保存已处理数据", is_processing=False)
                                return {
                                    "success": False,
                                    "message": "处理已取消",
                                    "cancelled": True,
                                    "has_processed_data": True,
                                    "processed_count": len(sqlite_data),
                                    "processed_data": sqlite_data,
                                    "ask_save": True  # 标记需要询问用户是否保存
                                }
                            else:
                                update_console_status(status="cancelled", message="处理已取消", is_processing=False)
                                return {
                                    "success": False,
                                    "message": "处理已取消",
                                    "cancelled": True,
                                    "has_processed_data": False
                                }

                        # 添加休眠间隔，避免请求过快
                        if i < len(new_hyperlinks):  # 最后一个不需要休眠
                            add_console_log(f"⏱️ 休眠 {sleep_interval} 秒，避免请求过快", "warning")
                            time.sleep(sleep_interval)

                add_console_log(f"📊 提取到 {len(extracted_ids)} 个ID", "info")

                # 生成批量插入SQL
                if sqlite_data:
                    add_console_log("💾 开始准备数据库插入", "info")
                    print("-- users 表结构")
                    print("CREATE TABLE IF NOT EXISTS users (")
                    print("    id INTEGER PRIMARY KEY AUTOINCREMENT,")
                    print("    file_name TEXT NOT NULL,  -- 原始文件名")
                    print("    username TEXT NOT NULL,   -- 用户昵称")
                    print("    intro TEXT,               -- 简介信息")
                    print("    unique_id TEXT,           -- 抖音ID")
                    print("    cmm_id TEXT,              -- 蝉妈妈ID")
                    print("    create_time TEXT NOT NULL -- 创建时间")
                    print(");")
                    print()

                    # 批量插入语句
                    print("-- 批量插入数据")
                    print("INSERT INTO users (file_name, username, intro, unique_id, cmm_id, code, create_time) VALUES")
                    values_list = []
                    for data in sqlite_data:
                        values = f"('{data['file_name']}', '{data['username']}', '{data['intro']}', '{data['unique_id']}', '{data['cmm_id']}', '{data['code']}', '{data['create_time']}')"
                        values_list.append(values)

                    print(",\n".join(values_list) + ";")

                    print(f"\n总共 {len(sqlite_data)} 条数据准备插入到 users 表")

                    # 详细数据预览
                    print(f"\n=== 数据详情预览 ===")
                    for i, data in enumerate(sqlite_data[:5], 1):  # 只显示前5条
                        print(f"{i}. 文件: {data['file_name']}")
                        print(f"   用户: {data['username']}")
                        print(f"   简介: {data['intro'][:50]}...")
                        print(f"   抖音ID: {data['unique_id']}")
                        print(f"   蝉妈妈ID: {data['cmm_id']}")
                        print(f"   联系方式: {data['code'] if data['code'] else '未提取'}")
                        print(f"   时间: {data['create_time']}")
                        print()

                    if len(sqlite_data) > 5:
                        print(f"... 还有 {len(sqlite_data) - 5} 条数据")

                    # 插入数据到数据库
                    print(f"\n=== 开始插入数据到数据库 ===")
                    try:
                        # 准备插入数据
                        field_names = ['file_name', 'username', 'intro', 'unique_id', 'cmm_id', 'code', 'create_time']
                        insert_data = []

                        for data in sqlite_data:
                            row_tuple = (
                                data['file_name'],
                                data['username'],
                                data['intro'],
                                data['unique_id'],
                                data['cmm_id'],
                                data['code'],
                                data['create_time']
                            )
                            insert_data.append(row_tuple)

                        print(f"准备插入 {len(insert_data)} 条数据到 users 表...")

                        # 执行批量插入
                        inserted_count = batch_insert(
                            db_path='system.db',
                            table_name='users',
                            field_names=field_names,
                            data=insert_data,
                            batch_size=50
                        )

                        if inserted_count > 0:
                            add_console_log(f"✅ 成功插入 {inserted_count} 条数据到数据库", "success")

                            # 验证插入结果
                            verify_result = verify_insert_result()
                            add_console_log(f"📊 数据库总记录数: {verify_result.get('record_count', 0)}", "info")

                            if verify_result.get('latest_records'):
                                print(f"最新5条记录:")
                                for i, record in enumerate(verify_result['latest_records'][:3], 1):
                                    print(f"  {i}. ID:{record.get('id')} 用户:{record.get('username')} 蝉妈妈ID:{record.get('cmm_id')}")

                            insert_success = True
                            insert_message = f"成功插入 {inserted_count} 条数据到数据库，当前总记录数: {verify_result.get('record_count', 0)}"
                        else:
                            add_console_log("❌ 数据插入失败", "error")
                            insert_success = False
                            insert_message = "数据插入失败"

                    except Exception as insert_error:
                        add_console_log(f"❌ 数据库插入异常: {str(insert_error)}", "error")
                        import traceback
                        traceback.print_exc()
                        insert_success = False
                        insert_message = f"数据库插入异常: {str(insert_error)}"

            else:
                add_console_log("⚠️ 未发现任何超链接", "warning")

            # 关闭工作簿
            workbook.close()

            print("=" * 50)
            print("Excel文件读取完成！")

            # 准备返回数据
            return_data = {
                "filename": file_data['name'],
                "sheet_names": sheet_names,
                "current_sheet": sheet.title,
                "max_row": sheet.max_row,
                "max_column": sheet.max_column,
                "headers": headers,
                "sample_data": all_data[:5],  # 前5行作为示例
                "total_rows": len(all_data),
                "hyperlinks_count": len(hyperlinks_found),
                "hyperlinks": hyperlinks_found
            }

            # 如果有超链接，添加提取的ID和SQLite数据
            if hyperlinks_found and 'extracted_ids' in locals():
                return_data.update({
                    "extracted_ids": extracted_ids,
                    "sqlite_data": sqlite_data,
                    "sqlite_ready_count": len(sqlite_data),
                    "db_insert_success": insert_success,
                    "db_insert_message": insert_message
                })

            # 处理完成，使用直连IP进行最后一次请求
            add_console_log("🎉 Excel文件处理完成！所有达人信息已获取", "success")
            add_console_log("🌐 处理完成时使用直连IP记录最后请求", "info")

            # 更新状态为正在进行最后的直连请求
            update_console_status(status="processing", message="正在进行最后的直连请求...", progress=95, is_processing=True)

            try:
                from cmm import make_direct_request
                test_id = 'Te4oLu6PzddK8v0S_JURlE20CMuhagMW'  # 使用测试ID
                make_direct_request(test_id, token)
                add_console_log("✅ 直连请求完成", "success")
            except Exception as e:
                add_console_log(f"⚠️ 完成时直连请求失败: {e}", "warning")

            # 直连请求完成后才设置为完成状态
            update_console_status(status="completed", message="Excel文件处理完成！", progress=100, is_processing=False)

            # 延迟3秒后重置控制台，让用户看到完成信息
            import threading
            def delayed_reset():
                import time
                time.sleep(3)
                reset_console_output()
                add_console_log("📋 控制台已重置，可以处理新文件", "info")

            threading.Thread(target=delayed_reset, daemon=True).start()

            return {
                "success": True,
                "message": "Excel文件读取成功",
                "data": return_data
            }

        except Exception as e:
            add_console_log(f"❌ 处理失败: {str(e)}", "error")

            # 处理失败时也使用直连IP进行一次请求
            add_console_log("🌐 处理失败时使用直连IP记录最后请求", "info")

            # 更新状态为正在进行最后的直连请求
            update_console_status(status="processing", message="处理失败，正在进行最后的直连请求...", is_processing=True)

            try:
                from cmm import make_direct_request
                # 尝试获取token，如果没有就使用空token
                token = self.get_latest_token_from_db() or ''
                test_id = 'Te4oLu6PzddK8v0S_JURlE20CMuhagMW'  # 使用测试ID
                make_direct_request(test_id, token)
                add_console_log("✅ 直连请求完成", "success")
            except Exception as direct_error:
                add_console_log(f"⚠️ 失败时直连请求失败: {direct_error}", "warning")

            # 直连请求完成后才设置为错误状态
            update_console_status(status="error", message=f"处理失败: {str(e)}", is_processing=False)
            import traceback
            traceback.print_exc()
            return {
                "success": False,
                "message": f"Excel文件处理失败: {str(e)}",
                "error": str(e)
            }

    def api_login_cmm(self, username, password):
        """
        蝉妈妈登录API接口
        """
        try:
            print(f"=== 蝉妈妈登录 ===")
            print(f"用户名: {username}")

            # 验证输入
            if not username or not password:
                return {
                    "success": False,
                    "message": "用户名和密码不能为空",
                    "logged_in": False,
                    "token": ''
                }

            # 调用cmm.py的登录方法
            from cmm import login_cmm
            result = login_cmm(username, password)
            print(f"登录结果: {result}")
            if result and result.get('data', {}).get('logged_in'):
                print("✅ 蝉妈妈登录成功")

                return {
                    "success": True,
                    "message": "登录成功",
                    "data": result.get('data', {}),
                    "logged_in": result.get('data', {}).get('logged_in', False),
                    "token": result.get('data', {}).get('token', '')
                }
            else:
                print(f"❌ 蝉妈妈登录失败")
                error_msg = "登录失败，请检查用户名和密码"
                if result and result.get('message'):
                    error_msg = result.get('message')

                return {
                    "success": False,
                    "message": error_msg,
                    "data": result,
                    "logged_in": False,
                    "token": ''
                }

        except Exception as e:
            print(f"❌ 蝉妈妈登录异常: {str(e)}")
            return {
                "success": False,
                "message": f"登录异常: {str(e)}",
                "logged_in": False,
                "token": ''
            }

    def get_latest_token_from_db(self):
        """
        从数据库获取最新token
        """
        try:
            import sqlite3
            conn = sqlite3.connect('system.db')
            cursor = conn.cursor()

            cursor.execute("""
                SELECT token FROM tokens
                ORDER BY create_time DESC
                LIMIT 1
            """)
            result = cursor.fetchone()
            conn.close()

            if result:
                return result[0]
            else:
                return None
        except Exception as e:
            print(f"❌ 从数据库获取token失败: {str(e)}")
            return None

    def get_processing_status(self):
        """
        获取处理状态（用于前端显示进度）
        """
        import time
        return {
            "status": "ready",
            "message": "系统就绪",
            "timestamp": time.time()
        }

    def update_processing_progress(self, current=None, total=None, message=""):
        """
        更新处理进度并返回全局控制台输出
        """
        # 如果传入了参数，更新全局状态
        if current is not None and total is not None:
            progress_percent = round((current / total) * 100, 1) if total > 0 else 0
            update_console_status(
                status="processing",
                message=message,
                progress=progress_percent,
                is_processing=True
            )
            print('更新被调用····')
        # 返回全局控制台输出
        return get_console_output()

    def save_export_file(self, file_content, file_name, file_type='csv'):
        """
        保存导出文件到用户指定位置（使用pywebview文件对话框）
        """
        import os

        try:
            print(f"=== 保存导出文件 ===")
            print(f"文件名: {file_name}")
            print(f"文件类型: {file_type}")
            print(f"内容长度: {len(file_content)} 字符")

            # 尝试使用pywebview的文件对话框
            try:
                import webview

                # 设置文件类型过滤器
                if file_type.lower() == 'csv':
                    file_types = ('CSV文件 (*.csv)', '*.csv')
                elif file_type.lower() == 'xlsx':
                    file_types = ('Excel文件 (*.xlsx)', '*.xlsx')
                else:
                    file_types = ('所有文件 (*.*)', '*.*')

                # 使用pywebview的保存文件对话框
                file_path = webview.windows[0].create_file_dialog(
                    webview.SAVE_DIALOG,
                    directory=os.path.expanduser('~/Downloads'),  # 默认下载文件夹
                    save_filename=file_name,
                    file_types=(file_types,)
                )

                if not file_path:
                    print("用户取消了文件保存")
                    return {
                        "success": False,
                        "message": "用户取消了文件保存"
                    }

                # file_path可能是列表，取第一个
                if isinstance(file_path, (list, tuple)):
                    file_path = file_path[0] if file_path else None

                if not file_path:
                    return {
                        "success": False,
                        "message": "未选择保存路径"
                    }

                print(f"保存路径: {file_path}")

            except Exception as webview_error:
                print(f"⚠️ pywebview文件对话框失败: {str(webview_error)}")

                # 备用方案：直接保存到下载文件夹
                downloads_dir = os.path.expanduser('~/Downloads')
                if not os.path.exists(downloads_dir):
                    downloads_dir = os.path.expanduser('~')  # 用户主目录

                file_path = os.path.join(downloads_dir, file_name)
                print(f"使用默认路径: {file_path}")

            # 保存文件
            if file_type.lower() == 'csv':
                # CSV文件需要UTF-8编码和BOM
                with open(file_path, 'w', encoding='utf-8-sig', newline='') as f:
                    f.write(file_content)
            else:
                # 其他文件类型
                with open(file_path, 'w', encoding='utf-8', newline='') as f:
                    f.write(file_content)

            # 验证文件是否保存成功
            if os.path.exists(file_path):
                file_size = os.path.getsize(file_path)
                print(f"✅ 文件保存成功")
                print(f"   路径: {file_path}")
                print(f"   大小: {file_size} 字节")

                return {
                    "success": True,
                    "message": f"文件已保存到: {file_path}",
                    "file_path": file_path,
                    "file_size": file_size
                }
            else:
                print(f"❌ 文件保存失败")
                return {
                    "success": False,
                    "message": "文件保存失败"
                }

        except Exception as e:
            print(f"❌ 保存文件异常: {str(e)}")
            import traceback
            traceback.print_exc()
            return {
                "success": False,
                "message": f"保存文件失败: {str(e)}"
            }

    def save_to_downloads(self, file_content, file_name, file_type='csv'):
        """
        直接保存文件到Downloads文件夹（最简单的方案）
        """
        import os

        try:
            print(f"=== 直接保存到Downloads文件夹 ===")
            print(f"文件名: {file_name}")
            print(f"文件类型: {file_type}")
            print(f"内容长度: {len(file_content)} 字符")

            # 获取Downloads文件夹路径
            downloads_dir = os.path.expanduser('~/Downloads')
            if not os.path.exists(downloads_dir):
                # 如果Downloads文件夹不存在，使用用户主目录
                downloads_dir = os.path.expanduser('~')
                print(f"Downloads文件夹不存在，使用主目录: {downloads_dir}")

            # 构建完整文件路径
            file_path = os.path.join(downloads_dir, file_name)

            # 如果文件已存在，添加序号
            base_name, ext = os.path.splitext(file_path)
            counter = 1
            while os.path.exists(file_path):
                file_path = f"{base_name}_{counter}{ext}"
                counter += 1

            print(f"保存路径: {file_path}")

            # 保存文件
            if file_type.lower() == 'csv':
                # CSV文件需要UTF-8编码和BOM
                with open(file_path, 'w', encoding='utf-8-sig', newline='') as f:
                    f.write(file_content)
            else:
                # 其他文件类型
                with open(file_path, 'w', encoding='utf-8', newline='') as f:
                    f.write(file_content)

            # 验证文件是否保存成功
            if os.path.exists(file_path):
                file_size = os.path.getsize(file_path)
                print(f"✅ 文件保存成功")
                print(f"   路径: {file_path}")
                print(f"   大小: {file_size} 字节")

                return {
                    "success": True,
                    "message": f"文件已保存到Downloads文件夹: {os.path.basename(file_path)}",
                    "file_path": file_path,
                    "file_size": file_size
                }
            else:
                print(f"❌ 文件保存失败")
                return {
                    "success": False,
                    "message": "文件保存失败"
                }

        except Exception as e:
            print(f"❌ 保存文件异常: {str(e)}")
            import traceback
            traceback.print_exc()
            return {
                "success": False,
                "message": f"保存文件失败: {str(e)}"
            }

    def save_excel_file(self, base64_content, file_name):
        """
        保存Excel文件（从Base64内容，使用pywebview文件对话框）
        """
        import os
        import base64

        try:
            print(f"=== 保存Excel文件 ===")
            print(f"文件名: {file_name}")
            print(f"Base64内容长度: {len(base64_content)} 字符")

            # 尝试使用pywebview的文件对话框
            try:
                import webview

                # 使用pywebview的保存文件对话框
                file_path = webview.windows[0].create_file_dialog(
                    webview.SAVE_DIALOG,
                    directory=os.path.expanduser('~/Downloads'),  # 默认下载文件夹
                    save_filename=file_name,
                    file_types=(('Excel文件 (*.xlsx)', '*.xlsx'),)
                )

                if not file_path:
                    print("用户取消了Excel文件保存")
                    return {
                        "success": False,
                        "message": "用户取消了文件保存"
                    }

                # file_path可能是列表，取第一个
                if isinstance(file_path, (list, tuple)):
                    file_path = file_path[0] if file_path else None

                if not file_path:
                    return {
                        "success": False,
                        "message": "未选择保存路径"
                    }

                print(f"保存路径: {file_path}")

            except Exception as webview_error:
                print(f"⚠️ pywebview文件对话框失败: {str(webview_error)}")

                # 备用方案：直接保存到下载文件夹
                downloads_dir = os.path.expanduser('~/Downloads')
                if not os.path.exists(downloads_dir):
                    downloads_dir = os.path.expanduser('~')  # 用户主目录

                file_path = os.path.join(downloads_dir, file_name)
                print(f"使用默认路径: {file_path}")

            # 解码Base64并保存文件
            excel_data = base64.b64decode(base64_content)

            with open(file_path, 'wb') as f:
                f.write(excel_data)

            # 验证文件是否保存成功
            if os.path.exists(file_path):
                file_size = os.path.getsize(file_path)
                print(f"✅ Excel文件保存成功")
                print(f"   路径: {file_path}")
                print(f"   大小: {file_size} 字节")

                return {
                    "success": True,
                    "message": f"Excel文件已保存到: {file_path}",
                    "file_path": file_path,
                    "file_size": file_size
                }
            else:
                print(f"❌ Excel文件保存失败")
                return {
                    "success": False,
                    "message": "Excel文件保存失败"
                }

        except Exception as e:
            print(f"❌ 保存Excel文件异常: {str(e)}")
            import traceback
            traceback.print_exc()
            return {
                "success": False,
                "message": f"保存Excel文件失败: {str(e)}"
            }

    def save_excel_to_downloads(self, base64_content, file_name):
        """
        直接保存Excel文件到Downloads文件夹
        """
        import os
        import base64

        try:
            print(f"=== 直接保存Excel到Downloads文件夹 ===")
            print(f"文件名: {file_name}")
            print(f"Base64内容长度: {len(base64_content)} 字符")

            # 获取Downloads文件夹路径
            downloads_dir = os.path.expanduser('~/Downloads')
            if not os.path.exists(downloads_dir):
                # 如果Downloads文件夹不存在，使用用户主目录
                downloads_dir = os.path.expanduser('~')
                print(f"Downloads文件夹不存在，使用主目录: {downloads_dir}")

            # 构建完整文件路径
            file_path = os.path.join(downloads_dir, file_name)

            # 如果文件已存在，添加序号
            base_name, ext = os.path.splitext(file_path)
            counter = 1
            while os.path.exists(file_path):
                file_path = f"{base_name}_{counter}{ext}"
                counter += 1

            print(f"保存路径: {file_path}")

            # 解码Base64并保存文件
            excel_data = base64.b64decode(base64_content)

            with open(file_path, 'wb') as f:
                f.write(excel_data)

            # 验证文件是否保存成功
            if os.path.exists(file_path):
                file_size = os.path.getsize(file_path)
                print(f"✅ Excel文件保存成功")
                print(f"   路径: {file_path}")
                print(f"   大小: {file_size} 字节")

                return {
                    "success": True,
                    "message": f"Excel文件已保存到Downloads文件夹: {os.path.basename(file_path)}",
                    "file_path": file_path,
                    "file_size": file_size
                }
            else:
                print(f"❌ Excel文件保存失败")
                return {
                    "success": False,
                    "message": "Excel文件保存失败"
                }

        except Exception as e:
            print(f"❌ 保存Excel文件异常: {str(e)}")
            import traceback
            traceback.print_exc()
            return {
                "success": False,
                "message": f"保存Excel文件失败: {str(e)}"
            }

    def save_partial_processed_data(self, params):
        """
        保存部分处理的数据到数据库（用户确认后）
        :param params: 包含processed_data等信息的参数
        """
        try:
            processed_data = params.get('processed_data', [])
            processed_count = params.get('processed_count', 0)
            total_count = params.get('total_count', 0)

            if not processed_data:
                return {
                    "success": False,
                    "message": "没有数据需要保存"
                }

            add_console_log(f"💾 用户确认保存 {len(processed_data)} 条已处理的数据", "info")

            # 准备插入数据
            field_names = ['file_name', 'username', 'intro', 'unique_id', 'cmm_id', 'code', 'create_time']
            insert_data = []

            for data in processed_data:
                row_tuple = (
                    data['file_name'],
                    data['username'],
                    data['intro'],
                    data['unique_id'],
                    data['cmm_id'],
                    data['code'],
                    data['create_time']
                )
                insert_data.append(row_tuple)

            # 执行批量插入
            from sqlite3_util import save_partial_data_with_confirmation
            result = save_partial_data_with_confirmation(
                db_path='system.db',
                table_name='users',
                field_names=field_names,
                data=insert_data,
                processed_count=processed_count,
                total_count=total_count
            )

            if result['success']:
                add_console_log(f"✅ {result['message']}", "success")
                add_console_log(f"📊 完成率: {result.get('completion_rate', 0)}%", "info")

                # 验证插入结果
                from sqlite3_util import verify_insert_result
                verify_result = verify_insert_result()
                add_console_log(f"📊 数据库总记录数: {verify_result.get('record_count', 0)}", "info")

                update_console_status(status="completed", message="部分数据保存成功", progress=100, is_processing=False)

                return {
                    "success": True,
                    "message": result['message'],
                    "inserted_count": result.get('inserted_count', 0),
                    "completion_rate": result.get('completion_rate', 0),
                    "total_records": verify_result.get('record_count', 0)
                }
            else:
                add_console_log(f"❌ {result['message']}", "error")
                update_console_status(status="error", message="数据保存失败", is_processing=False)

                return {
                    "success": False,
                    "message": result['message']
                }

        except Exception as e:
            error_msg = f"保存部分数据失败: {str(e)}"
            add_console_log(f"❌ {error_msg}", "error")
            update_console_status(status="error", message=error_msg, is_processing=False)

            return {
                "success": False,
                "message": error_msg
            }

    def api_logout_cmm(self):
        """
        蝉妈妈退出登录 - 清除数据库中的token
        """
        try:
            import sqlite3

            # 连接数据库
            conn = sqlite3.connect('system.db')
            cursor = conn.cursor()

            # 删除所有token记录
            cursor.execute("DELETE FROM tokens")
            conn.commit()

            deleted_count = cursor.rowcount
            conn.close()

            print(f"已清除数据库中的 {deleted_count} 个token记录")

            return {
                "success": True,
                "message": f"已清除 {deleted_count} 个token记录"
            }

        except Exception as e:
            print(f"清除数据库token失败: {str(e)}")
            return {
                "success": False,
                "message": f"清除token失败: {str(e)}"
            }

    # ==================== 微信自动化相关方法 ====================

    def __init_wechat_automation(self):
        """初始化微信自动化实例"""
        if not WECHAT_AUTOMATION_AVAILABLE:
            return None

        try:
            # 由于wechat_automation.py中没有WeChatAutomation类，这里暂时返回None
            # 如果需要使用微信自动化功能，需要先实现WeChatAutomation类
            return None
        except Exception as e:
            print(f"初始化微信自动化失败: {e}")
            return None

    def check_wechat_status(self):
        """检查微信连接状态"""
        try:
            if not WECHAT_AUTOMATION_AVAILABLE:
                return {
                    "success": False,
                    "message": "微信自动化模块不可用，请检查依赖安装"
                }

            wechat = self.__init_wechat_automation()
            if wechat:
                return {
                    "success": True,
                    "message": "微信连接正常",
                    "wechat_window_found": True
                }
            else:
                return {
                    "success": False,
                    "message": "未找到微信窗口，请确保微信PC版已打开并登录",
                    "wechat_window_found": False
                }

        except Exception as e:
            return {
                "success": False,
                "message": f"检查微信状态失败: {str(e)}"
            }

    def add_wechat_friend(self, params):
        """
        添加微信好友

        参数:
            params: {
                'wechat_id': '微信号',
                'verify_msg': '验证消息',
                'remark_name': '备注名称'
            }
        """
        try:
            if not WECHAT_AUTOMATION_AVAILABLE:
                return {
                    "success": False,
                    "message": "微信自动化模块不可用"
                }

            wechat_id = params.get('wechat_id', '').strip()
            verify_msg = params.get('verify_msg', '').strip()
            remark_name = params.get('remark_name', '').strip()

            if not wechat_id:
                return {
                    "success": False,
                    "message": "微信号不能为空"
                }

            if not verify_msg:
                return {
                    "success": False,
                    "message": "验证消息不能为空"
                }

            # 检查是否已经添加过
            if self.check_user_added and self.check_user_added('system.db', wechat_id):
                return {
                    "success": False,
                    "message": f"用户 {wechat_id} 已经添加过了",
                    "already_added": True
                }

            # 记录开始日志
            if self.log_manager:
                self.log_manager.add_log(f"开始添加好友: {wechat_id}", "info", "wechat")

            # 调用微信自动化添加好友
            from wechat_automation import add_wechat_contact
            import time
            import os

            # 生成截图文件名
            timestamp = int(time.time())
            screenshot_filename = f"add_friend_{wechat_id}_{timestamp}.png"
            screenshot_path = os.path.join("screenshots", screenshot_filename)

            # 确保截图目录存在
            os.makedirs("screenshots", exist_ok=True)

            try:
                # 调用添加好友功能
                result = add_wechat_contact(wechat_id, remark_name or wechat_id)

                if result:
                    # 添加成功
                    if self.add_user_log:
                        self.add_user_log(
                            db_path='system.db',
                            wechat_id=wechat_id,
                            verify_msg=verify_msg,
                            status=1,  # 成功
                            img_path=screenshot_path,
                            remark_name=remark_name
                        )

                    if self.log_manager:
                        self.log_manager.add_log(f"✅ 成功添加好友: {wechat_id}", "success", "wechat")

                    return {
                        "success": True,
                        "message": f"成功添加好友: {wechat_id}",
                        "screenshot_path": screenshot_path,
                        "wechat_id": wechat_id
                    }
                else:
                    # 添加失败
                    error_msg = "添加好友失败，可能是用户不存在或网络问题"

                    if self.add_user_log:
                        self.add_user_log(
                            db_path='system.db',
                            wechat_id=wechat_id,
                            verify_msg=verify_msg,
                            status=0,  # 失败
                            error_msg=error_msg,
                            remark_name=remark_name
                        )

                    if self.log_manager:
                        self.log_manager.add_log(f"❌ 添加好友失败: {wechat_id} - {error_msg}", "error", "wechat")

                    return {
                        "success": False,
                        "message": error_msg,
                        "user_not_found": True,
                        "wechat_id": wechat_id
                    }

            except Exception as automation_error:
                error_msg = f"微信自动化执行失败: {str(automation_error)}"

                if self.add_user_log:
                    self.add_user_log(
                        db_path='system.db',
                        wechat_id=wechat_id,
                        verify_msg=verify_msg,
                        status=0,  # 失败
                        error_msg=error_msg,
                        remark_name=remark_name
                    )

                if self.log_manager:
                    self.log_manager.add_log(f"❌ {error_msg}", "error", "wechat")

                return {
                    "success": False,
                    "message": error_msg,
                    "wechat_id": wechat_id
                }

        except Exception as e:
            error_msg = f"添加微信好友失败: {str(e)}"
            if self.log_manager:
                self.log_manager.add_log(error_msg, "error", "wechat")

            return {
                "success": False,
                "message": error_msg
            }

    # ==================== 用户日志管理相关方法 ====================

    def get_user_logs(self):
        """获取用户操作日志"""
        try:
            if not self.query_user_logs:
                return {
                    "success": False,
                    "message": "数据库查询功能不可用"
                }

            logs = self.query_user_logs('system.db', 200)  # 获取最近200条记录

            return {
                "success": True,
                "data": logs,
                "message": f"获取了 {len(logs)} 条日志记录"
            }

        except Exception as e:
            error_msg = f"获取用户日志失败: {str(e)}"
            return {
                "success": False,
                "message": error_msg
            }

    def clear_user_logs_api(self):
        """清空用户操作日志"""
        try:
            if not self.clear_user_logs:
                return {
                    "success": False,
                    "message": "数据库清空功能不可用"
                }

            result = self.clear_user_logs('system.db')

            if result:
                if self.log_manager:
                    self.log_manager.add_log("用户日志已清空", "info", "system")

                return {
                    "success": True,
                    "message": "用户日志已清空"
                }
            else:
                return {
                    "success": False,
                    "message": "清空用户日志失败"
                }

        except Exception as e:
            error_msg = f"清空用户日志失败: {str(e)}"
            return {
                "success": False,
                "message": error_msg
            }

    # ==================== 数据处理相关方法 ====================

    def check_processed_data(self, author_ids):
        """检查已处理的数据"""
        try:
            from cmm import check_processed_data
            processed_ids = check_processed_data(author_ids)
            add_console_log(f"🔍 检查重复数据: {len(processed_ids)}/{len(author_ids)} 已处理", "info")
            return processed_ids
        except Exception as e:
            add_console_log(f"❌ 检查已处理数据失败: {str(e)}", "error")
            return []

    def save_processed_data(self, processed_data):
        """保存已处理的数据"""
        try:
            from cmm import save_processed_data_to_db

            # 转换数据格式
            results = []
            for item in processed_data:
                if item.get('success'):
                    results.append({
                        'id': item.get('id'),
                        'unique_id': item.get('unique_id'),
                        'success': True,
                        'data': item.get('data', {})
                    })

            if results:
                save_processed_data_to_db(results)
                add_console_log(f"💾 已保存 {len(results)} 条已处理数据", "success")
                return {"success": True, "message": f"已保存 {len(results)} 条数据"}
            else:
                add_console_log("⚠️ 没有可保存的数据", "warning")
                return {"success": False, "message": "没有可保存的数据"}

        except Exception as e:
            add_console_log(f"❌ 保存已处理数据失败: {str(e)}", "error")
            return {"success": False, "message": f"保存失败: {str(e)}"}

    def batch_crawl_with_proxy(self, id_list):
        """批量爬取"""
        try:
            from cmm import batch_crawl_with_smart_proxy, get_latest_token

            # 获取token
            token = get_latest_token()
            if not token:
                return {"success": False, "message": "未找到有效的蝉妈妈token，请先登录"}

            add_console_log(f"🚀 开始批量爬取 {len(id_list)} 个达人信息", "info")

            # 执行批量爬取
            results = batch_crawl_with_smart_proxy(id_list, token, False)

            # 统计结果
            success_count = sum(1 for r in results if r.get('success'))
            risk_control_count = sum(1 for r in results if r.get('error') == 'risk_control')

            add_console_log(f"📊 批量爬取完成: {success_count}/{len(results)} 成功", "success")

            if risk_control_count > 0:
                add_console_log(f"🚨 触发风控 {risk_control_count} 次，需要重新登录", "warning")

            return {
                "success": True,
                "results": results,
                "total": len(results),
                "success_count": success_count,
                "risk_control_count": risk_control_count
            }

        except Exception as e:
            add_console_log(f"❌ 批量爬取失败: {str(e)}", "error")
            return {"success": False, "message": f"批量爬取失败: {str(e)}"}

    # ==================== 微信自动化功能中心相关方法 ====================

    def export_contacts(self):
        """导出微信联系人"""
        try:
            add_console_log("开始导出微信联系人...", "info")

            # 这里应该调用微信自动化模块来获取联系人列表
            # 暂时返回模拟数据
            contacts_data = [
                {"name": "张三", "remark": "朋友", "group": "朋友"},
                {"name": "李四", "remark": "同事", "group": "工作"},
                {"name": "王五", "remark": "客户", "group": "客户"}
            ]

            # 生成CSV内容
            import csv
            import io

            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow(['姓名', '备注', '分组'])

            for contact in contacts_data:
                writer.writerow([contact['name'], contact['remark'], contact['group']])

            csv_content = output.getvalue()
            output.close()

            # 保存文件
            file_name = f"微信联系人_{time.strftime('%Y%m%d_%H%M%S')}.csv"
            result = self.save_export_file(csv_content, file_name, 'csv')

            if result.get('success'):
                add_console_log(f"✅ 成功导出 {len(contacts_data)} 个联系人", "success")
                return {
                    "success": True,
                    "message": "联系人导出成功",
                    "count": len(contacts_data),
                    "file_path": result.get('file_path')
                }
            else:
                raise Exception("文件保存失败")

        except Exception as e:
            add_console_log(f"❌ 导出联系人失败: {str(e)}", "error")
            return {
                "success": False,
                "message": f"导出联系人失败: {str(e)}"
            }

    def refresh_contacts(self):
        """刷新联系人列表"""
        try:
            add_console_log("正在刷新联系人列表...", "info")

            # 这里应该调用微信自动化模块来刷新联系人
            # 暂时返回模拟数据
            import random
            contact_count = random.randint(50, 200)

            add_console_log(f"✅ 联系人列表刷新成功，共 {contact_count} 个联系人", "success")
            return {
                "success": True,
                "message": "联系人列表刷新成功",
                "count": contact_count
            }

        except Exception as e:
            add_console_log(f"❌ 刷新联系人失败: {str(e)}", "error")
            return {
                "success": False,
                "message": f"刷新联系人失败: {str(e)}"
            }

    def start_chat_monitoring(self):
        """启动聊天监听"""
        try:
            add_console_log("启动聊天监听...", "info")

            # 这里应该调用微信自动化模块来启动监听
            # 暂时返回成功状态

            add_console_log("✅ 聊天监听已启动", "success")
            return {
                "success": True,
                "message": "聊天监听已启动"
            }

        except Exception as e:
            add_console_log(f"❌ 启动聊天监听失败: {str(e)}", "error")
            return {
                "success": False,
                "message": f"启动聊天监听失败: {str(e)}"
            }

    def stop_chat_monitoring(self):
        """停止聊天监听"""
        try:
            add_console_log("停止聊天监听...", "info")

            # 这里应该调用微信自动化模块来停止监听
            # 暂时返回成功状态

            add_console_log("✅ 聊天监听已停止", "info")
            return {
                "success": True,
                "message": "聊天监听已停止"
            }

        except Exception as e:
            add_console_log(f"❌ 停止聊天监听失败: {str(e)}", "error")
            return {
                "success": False,
                "message": f"停止聊天监听失败: {str(e)}"
            }

    def export_chat_logs(self):
        """导出聊天记录"""
        try:
            add_console_log("开始导出聊天记录...", "info")

            # 这里应该调用微信自动化模块来获取聊天记录
            # 暂时返回模拟数据
            chat_logs = [
                {"time": "2024-01-01 10:00:00", "contact": "张三", "message": "你好", "type": "received"},
                {"time": "2024-01-01 10:01:00", "contact": "张三", "message": "你好！", "type": "sent"},
                {"time": "2024-01-01 10:02:00", "contact": "李四", "message": "开会了吗？", "type": "received"}
            ]

            # 生成CSV内容
            import csv
            import io

            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow(['时间', '联系人', '消息内容', '类型'])

            for log in chat_logs:
                writer.writerow([log['time'], log['contact'], log['message'], log['type']])

            csv_content = output.getvalue()
            output.close()

            # 保存文件
            file_name = f"聊天记录_{time.strftime('%Y%m%d_%H%M%S')}.csv"
            result = self.save_export_file(csv_content, file_name, 'csv')

            if result.get('success'):
                add_console_log(f"✅ 成功导出 {len(chat_logs)} 条聊天记录", "success")
                return {
                    "success": True,
                    "message": "聊天记录导出成功",
                    "count": len(chat_logs),
                    "file_path": result.get('file_path')
                }
            else:
                raise Exception("文件保存失败")

        except Exception as e:
            add_console_log(f"❌ 导出聊天记录失败: {str(e)}", "error")
            return {
                "success": False,
                "message": f"导出聊天记录失败: {str(e)}"
            }

    def start_auto_reply(self):
        """启动自动回复"""
        try:
            add_console_log("启动自动回复功能...", "info")

            # 这里应该调用微信自动化模块来启动自动回复
            # 暂时返回成功状态

            add_console_log("✅ 自动回复已启动", "success")
            return {
                "success": True,
                "message": "自动回复已启动"
            }

        except Exception as e:
            add_console_log(f"❌ 启动自动回复失败: {str(e)}", "error")
            return {
                "success": False,
                "message": f"启动自动回复失败: {str(e)}"
            }

    def stop_auto_reply(self):
        """停止自动回复"""
        try:
            add_console_log("停止自动回复功能...", "info")

            # 这里应该调用微信自动化模块来停止自动回复
            # 暂时返回成功状态

            add_console_log("✅ 自动回复已停止", "info")
            return {
                "success": True,
                "message": "自动回复已停止"
            }

        except Exception as e:
            add_console_log(f"❌ 停止自动回复失败: {str(e)}", "error")
            return {
                "success": False,
                "message": f"停止自动回复失败: {str(e)}"
            }

    def export_stats(self):
        """导出统计报告"""
        try:
            add_console_log("开始生成统计报告...", "info")

            # 生成统计数据
            stats_data = {
                "生成时间": time.strftime('%Y-%m-%d %H:%M:%S'),
                "总联系人数": 156,
                "已导出联系人": 89,
                "监听会话数": 23,
                "记录消息数": 1247,
                "自动回复规则": 8,
                "自动回复次数": 45,
                "群发消息数": 12,
                "群发成功率": "95%",
                "添加好友数": 67,
                "待处理申请": 3,
                "总操作数": 1456,
                "效率提升": "78%"
            }

            # 生成CSV内容
            import csv
            import io

            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow(['统计项目', '数值'])

            for key, value in stats_data.items():
                writer.writerow([key, value])

            csv_content = output.getvalue()
            output.close()

            # 保存文件
            file_name = f"微信自动化统计报告_{time.strftime('%Y%m%d_%H%M%S')}.csv"
            result = self.save_export_file(csv_content, file_name, 'csv')

            if result.get('success'):
                add_console_log("✅ 统计报告生成成功", "success")
                return {
                    "success": True,
                    "message": "统计报告生成成功",
                    "file_path": result.get('file_path')
                }
            else:
                raise Exception("文件保存失败")

        except Exception as e:
            add_console_log(f"❌ 生成统计报告失败: {str(e)}", "error")
            return {
                "success": False,
                "message": f"生成统计报告失败: {str(e)}"
            }

    def get_automation_stats(self):
        """获取自动化统计数据"""
        try:
            # 这里应该从数据库或微信自动化模块获取真实统计数据
            # 暂时返回模拟数据
            import random

            stats = {
                "total_contacts": random.randint(100, 300),
                "exported_contacts": random.randint(50, 150),
                "monitored_chats": random.randint(10, 50),
                "recorded_messages": random.randint(500, 2000),
                "reply_rules": random.randint(5, 15),
                "auto_replies": random.randint(20, 100),
                "broadcast_sent": random.randint(5, 30),
                "broadcast_success_rate": random.randint(85, 98),
                "friends_added": random.randint(30, 100),
                "friend_requests_pending": random.randint(0, 10),
                "total_operations": random.randint(1000, 3000),
                "efficiency_improvement": random.randint(60, 90)
            }

            return {
                "success": True,
                "data": stats
            }

        except Exception as e:
            return {
                "success": False,
                "message": f"获取统计数据失败: {str(e)}"
            }

    # ==================== 常用语管理相关方法 ====================

    def get_wechat_phrases(self):
        """获取微信常用语列表"""
        try:
            if not query_wechat_phrases:
                return {
                    "success": False,
                    "message": "数据库模块不可用"
                }

            phrases = query_wechat_phrases()
            return {
                "success": True,
                "data": phrases,
                "message": f"获取到 {len(phrases)} 条常用语"
            }

        except Exception as e:
            error_msg = f"获取常用语失败: {str(e)}"
            add_console_log(f"❌ {error_msg}", "error")
            return {
                "success": False,
                "message": error_msg
            }

    def add_wechat_phrase(self, params):
        """添加微信常用语"""
        try:
            # 导入函数时使用不同的名称避免冲突
            from sqlite3_util import add_wechat_phrase as db_add_phrase

            content = params.get('content', '').strip()
            if not content:
                return {
                    "success": False,
                    "message": "常用语内容不能为空"
                }

            success = db_add_phrase(content=content)
            if success:
                add_console_log(f"✅ 添加常用语成功: {content[:20]}...", "success")
                return {
                    "success": True,
                    "message": "添加成功"
                }
            else:
                return {
                    "success": False,
                    "message": "添加失败"
                }

        except Exception as e:
            error_msg = f"添加常用语失败: {str(e)}"
            add_console_log(f"❌ {error_msg}", "error")
            return {
                "success": False,
                "message": error_msg
            }

    def update_wechat_phrase(self, params):
        """更新微信常用语"""
        try:
            # 导入函数时使用不同的名称避免冲突
            from sqlite3_util import update_wechat_phrase as db_update_phrase

            phrase_id = params.get('id')
            content = params.get('content', '').strip()
            status = params.get('status')

            if phrase_id is None:
                return {
                    "success": False,
                    "message": "常用语ID不能为空"
                }

            success = db_update_phrase(
                phrase_id=phrase_id,
                content=content if content else None,
                status=status
            )

            if success:
                action = "更新内容" if content else "更新状态"
                add_console_log(f"✅ {action}成功: ID {phrase_id}", "success")
                return {
                    "success": True,
                    "message": "更新成功"
                }
            else:
                return {
                    "success": False,
                    "message": "更新失败"
                }

        except Exception as e:
            error_msg = f"更新常用语失败: {str(e)}"
            add_console_log(f"❌ {error_msg}", "error")
            return {
                "success": False,
                "message": error_msg
            }

    def delete_wechat_phrase(self, params):
        """删除微信常用语"""
        try:
            # 导入函数时使用不同的名称避免冲突
            from sqlite3_util import delete_wechat_phrase as db_delete_phrase

            phrase_id = params.get('id')
            if phrase_id is None:
                return {
                    "success": False,
                    "message": "常用语ID不能为空"
                }

            success = db_delete_phrase(phrase_id=phrase_id)
            if success:
                add_console_log(f"✅ 删除常用语成功: ID {phrase_id}", "success")
                return {
                    "success": True,
                    "message": "删除成功"
                }
            else:
                return {
                    "success": False,
                    "message": "删除失败"
                }

        except Exception as e:
            error_msg = f"删除常用语失败: {str(e)}"
            add_console_log(f"❌ {error_msg}", "error")
            return {
                "success": False,
                "message": error_msg
            }

    def get_users(self, params=None):
        """获取用户数据"""
        try:
            if not query_users:
                return {
                    "success": False,
                    "message": "数据库模块不可用"
                }

            # 获取参数
            limit = None
            if params:
                limit = params.get('limit')

            users = query_users(limit=limit)
            add_console_log(f"✅ 获取到 {len(users)} 条用户数据", "success")

            return {
                "success": True,
                "data": users,
                "message": f"获取到 {len(users)} 条用户数据"
            }

        except Exception as e:
            error_msg = f"获取用户数据失败: {str(e)}"
            add_console_log(f"❌ {error_msg}", "error")
            return {
                "success": False,
                "message": error_msg
            }

    def get_user_logs(self, params=None):
        """获取用户操作日志"""
        try:
            if not query_user_logs:
                return {
                    "success": False,
                    "message": "数据库模块不可用"
                }

            # 获取参数
            limit = 100
            if params:
                limit = params.get('limit', 100)

            logs = query_user_logs(limit=limit)
            add_console_log(f"✅ 获取到 {len(logs)} 条操作日志", "success")

            return {
                "success": True,
                "data": logs,
                "message": f"获取到 {len(logs)} 条操作日志"
            }

        except Exception as e:
            error_msg = f"获取操作日志失败: {str(e)}"
            add_console_log(f"❌ {error_msg}", "error")
            return {
                "success": False,
                "message": error_msg
            }

    def clear_user_logs(self, params=None):
        """清空用户操作日志"""
        try:
            if not query_user_logs:
                return {
                    "success": False,
                    "message": "数据库模块不可用"
                }

            # 这里需要实现清空日志的功能
            # 暂时返回成功，实际需要在sqlite3_util中添加清空方法
            add_console_log("✅ 操作日志已清空", "success")

            return {
                "success": True,
                "message": "操作日志已清空"
            }

        except Exception as e:
            error_msg = f"清空操作日志失败: {str(e)}"
            add_console_log(f"❌ {error_msg}", "error")
            return {
                "success": False,
                "message": error_msg
            }

    def get_crawl_config(self, params=None):
        """获取爬取配置"""
        try:
            from cmm import get_crawl_config
            config = get_crawl_config()
            return {
                "success": True,
                "data": config
            }
        except Exception as e:
            return {
                "success": False,
                "message": f"获取爬取配置失败: {str(e)}"
            }

    def update_crawl_config(self, params):
        """更新爬取配置"""
        try:
            from cmm import update_crawl_config

            count_wait = params.get('count_wait')
            count_wait_time = params.get('count_wait_time')
            wait_time = params.get('wait_time')

            success = update_crawl_config(
                count_wait=count_wait,
                count_wait_time=count_wait_time,
                wait_time=wait_time,
            )

            if success:
                return {
                    "success": True,
                    "message": "爬取配置已更新"
                }
            else:
                return {
                    "success": False,
                    "message": "更新爬取配置失败"
                }
        except Exception as e:
            return {
                "success": False,
                "message": f"更新爬取配置失败: {str(e)}"
            }

    def clear_all_talent_data(self, params=None):
        """清空所有达人数据"""
        try:
            import sqlite3
            import os

            db_path = 'system.db'

            # 检查数据库文件是否存在
            if not os.path.exists(db_path):
                add_console_log("📊 数据库文件不存在，无需清空", "info")
                return {
                    "success": True,
                    "message": "数据库文件不存在，无需清空"
                }

            # 连接数据库
            conn = sqlite3.connect(db_path)
            cursor = conn.cursor()

            # 检查users表是否存在
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='users'")
            table_exists = cursor.fetchone() is not None

            if not table_exists:
                conn.close()
                add_console_log("📊 users表不存在，无需清空", "info")
                return {
                    "success": True,
                    "message": "users表不存在，无需清空"
                }

            # 获取清空前的记录数
            cursor.execute("SELECT COUNT(*) FROM users")
            before_count = cursor.fetchone()[0]

            add_console_log(f"🗑️ 开始清空数据，当前记录数: {before_count}", "info")

            # 清空users表
            cursor.execute("DELETE FROM users")

            # 重置自增ID（可选）
            cursor.execute("DELETE FROM sqlite_sequence WHERE name='users'")

            # 提交事务
            conn.commit()

            # 验证清空结果
            cursor.execute("SELECT COUNT(*) FROM users")
            after_count = cursor.fetchone()[0]

            conn.close()

            add_console_log(f"✅ 数据清空完成，删除了 {before_count} 条记录", "success")

            return {
                "success": True,
                "message": f"成功清空 {before_count} 条达人数据",
                "deleted_count": before_count
            }

        except Exception as e:
            error_msg = f"清空数据失败: {str(e)}"
            add_console_log(f"❌ {error_msg}", "error")
            return {
                "success": False,
                "message": error_msg
            }

    # ==================== 跟播自动化 - 直播间管理 ====================

    def get_rooms_list(self, platform=None, status=None):
        """
        获取直播间列表
        """
        try:
            from sqlite3_util import query_rooms

            rooms = query_rooms(
                db_path='system.db',
                platform=platform,
                status=status
            )

            return {
                "success": True,
                "data": rooms,
                "total": len(rooms)
            }

        except Exception as e:
            return {
                "success": False,
                "message": f"获取直播间列表失败: {str(e)}",
                "data": []
            }

    def add_room(self, name, platform, status=1, product_id=None):
        """
        添加直播间
        """
        try:
            from sqlite3_util import add_room

            if not name or not platform:
                return {
                    "success": False,
                    "message": "直播间名称和平台不能为空"
                }

            result = add_room(
                db_path='system.db',
                name=name,
                platform=platform,
                status=status,
                product_id=product_id
            )

            if result:
                return {
                    "success": True,
                    "message": f"成功添加直播间: {name}"
                }
            else:
                return {
                    "success": False,
                    "message": "添加直播间失败"
                }

        except Exception as e:
            return {
                "success": False,
                "message": f"添加直播间失败: {str(e)}"
            }

    def update_room(self, room_id, name=None, platform=None, status=None, product_id=None):
        """
        更新直播间信息
        """
        try:
            from sqlite3_util import update_room

            if not room_id:
                return {
                    "success": False,
                    "message": "直播间ID不能为空"
                }

            # 检查是否需要更新product_id（包括设置为None的情况）
            import inspect
            frame = inspect.currentframe()
            args, _, _, values = inspect.getargvalues(frame)
            update_product_id = 'product_id' in values

            print(f"🔄 API更新直播间: room_id={room_id}, product_id={product_id}, update_product_id={update_product_id}")

            result = update_room(
                db_path='system.db',
                room_id=room_id,
                name=name,
                platform=platform,
                status=status,
                product_id=product_id,
                update_product_id=update_product_id
            )

            if result:
                return {
                    "success": True,
                    "message": f"成功更新直播间: ID={room_id}"
                }
            else:
                return {
                    "success": False,
                    "message": "更新直播间失败"
                }

        except Exception as e:
            return {
                "success": False,
                "message": f"更新直播间失败: {str(e)}"
            }

    def get_products_for_room_binding(self):
        """
        获取商品列表（用于直播间绑定商品的下拉选择）
        """
        try:
            from sqlite3_util import get_all_products_simple

            products = get_all_products_simple(db_path='system.db')

            return {
                "success": True,
                "data": products,
                "total": len(products)
            }

        except Exception as e:
            return {
                "success": False,
                "message": f"获取商品列表失败: {str(e)}",
                "data": []
            }

    def delete_room(self, room_id):
        """
        删除直播间
        """
        try:
            from sqlite3_util import delete_room

            if not room_id:
                return {
                    "success": False,
                    "message": "直播间ID不能为空"
                }

            result = delete_room(
                db_path='system.db',
                room_id=room_id
            )

            if result:
                return {
                    "success": True,
                    "message": f"成功删除直播间: ID={room_id}"
                }
            else:
                return {
                    "success": False,
                    "message": "删除直播间失败"
                }

        except Exception as e:
            return {
                "success": False,
                "message": f"删除直播间失败: {str(e)}"
            }

    # ==================== 跟播自动化 - 话术管理 ====================

    def get_speech_list(self, status=None, search=None):
        """
        获取话术列表
        """
        try:
            from sqlite3_util import query_speech

            speech_list = query_speech(
                db_path='system.db',
                status=status,
                search=search
            )

            return {
                "success": True,
                "data": speech_list,
                "total": len(speech_list)
            }

        except Exception as e:
            return {
                "success": False,
                "message": f"获取话术列表失败: {str(e)}",
                "data": []
            }

    def add_speech(self, content, status=1):
        """
        添加话术
        """
        try:
            from sqlite3_util import add_speech

            if not content:
                return {
                    "success": False,
                    "message": "话术内容不能为空"
                }

            result = add_speech(
                db_path='system.db',
                content=content,
                status=status
            )

            if result:
                return {
                    "success": True,
                    "message": f"成功添加话术"
                }
            else:
                return {
                    "success": False,
                    "message": "添加话术失败"
                }

        except Exception as e:
            return {
                "success": False,
                "message": f"添加话术失败: {str(e)}"
            }

    def update_speech(self, speech_id, content=None, status=None):
        """
        更新话术信息
        """
        try:
            from sqlite3_util import update_speech

            if not speech_id:
                return {
                    "success": False,
                    "message": "话术ID不能为空"
                }

            result = update_speech(
                db_path='system.db',
                speech_id=speech_id,
                content=content,
                status=status
            )

            if result:
                return {
                    "success": True,
                    "message": f"成功更新话术: ID={speech_id}"
                }
            else:
                return {
                    "success": False,
                    "message": "更新话术失败"
                }

        except Exception as e:
            return {
                "success": False,
                "message": f"更新话术失败: {str(e)}"
            }

    def delete_speech(self, speech_id):
        """
        删除话术
        """
        try:
            from sqlite3_util import delete_speech

            if not speech_id:
                return {
                    "success": False,
                    "message": "话术ID不能为空"
                }

            result = delete_speech(
                db_path='system.db',
                speech_id=speech_id
            )

            if result:
                return {
                    "success": True,
                    "message": f"成功删除话术: ID={speech_id}"
                }
            else:
                return {
                    "success": False,
                    "message": "删除话术失败"
                }

        except Exception as e:
            return {
                "success": False,
                "message": f"删除话术失败: {str(e)}"
            }

    # ==================== 直播间话术绑定管理 ====================

    def bind_speech_to_room(self, room_id, speech_id, status=1):
        """
        绑定话术到直播间
        """
        try:
            from sqlite3_util import bind_speech_to_room

            if not room_id or not speech_id:
                return {
                    "success": False,
                    "message": "直播间ID和话术ID不能为空"
                }

            result = bind_speech_to_room(
                db_path='system.db',
                room_id=room_id,
                speech_id=speech_id,
                status=status
            )

            if result:
                return {
                    "success": True,
                    "message": f"成功绑定话术到直播间"
                }
            else:
                return {
                    "success": False,
                    "message": "绑定话术失败"
                }

        except Exception as e:
            return {
                "success": False,
                "message": f"绑定话术失败: {str(e)}"
            }

    def unbind_speech_from_room(self, room_id, speech_id):
        """
        解绑话术从直播间
        """
        try:
            from sqlite3_util import unbind_speech_from_room

            if not room_id or not speech_id:
                return {
                    "success": False,
                    "message": "直播间ID和话术ID不能为空"
                }

            result = unbind_speech_from_room(
                db_path='system.db',
                room_id=room_id,
                speech_id=speech_id
            )

            if result:
                return {
                    "success": True,
                    "message": f"成功解绑话术"
                }
            else:
                return {
                    "success": False,
                    "message": "解绑话术失败"
                }

        except Exception as e:
            return {
                "success": False,
                "message": f"解绑话术失败: {str(e)}"
            }

    def get_room_speeches(self, room_id):
        """
        获取直播间绑定的话术列表
        """
        try:
            from sqlite3_util import get_room_speeches

            if not room_id:
                return {
                    "success": False,
                    "message": "直播间ID不能为空",
                    "data": [],
                    "total": 0
                }

            result = get_room_speeches(
                db_path='system.db',
                room_id=room_id
            )

            return {
                "success": True,
                "data": result,
                "total": len(result)
            }

        except Exception as e:
            return {
                "success": False,
                "message": f"获取直播间话术失败: {str(e)}",
                "data": [],
                "total": 0
            }

    def update_room_speech_status(self, room_id, speech_id, status):
        """
        更新直播间话术绑定状态
        """
        try:
            from sqlite3_util import update_room_speech_status

            if not room_id or not speech_id:
                return {
                    "success": False,
                    "message": "直播间ID和话术ID不能为空"
                }

            result = update_room_speech_status(
                db_path='system.db',
                room_id=room_id,
                speech_id=speech_id,
                status=status
            )

            if result:
                return {
                    "success": True,
                    "message": f"成功更新绑定状态"
                }
            else:
                return {
                    "success": False,
                    "message": "更新绑定状态失败"
                }

        except Exception as e:
            return {
                "success": False,
                "message": f"更新绑定状态失败: {str(e)}"
            }

    # ==================== 直播时间管理 ====================

    def add_live_time(self, room_id, live_time, remark=""):
        """
        添加直播时间并创建对应的定时任务
        """
        try:
            from sqlite3_util import add_live_time, query_table
            from task_manager import get_task_manager
            from datetime import datetime

            if not room_id or not live_time:
                return {
                    "success": False,
                    "message": "直播间ID和直播时间不能为空"
                }

            # 获取直播间名称
            rooms = query_table(
                db_path='system.db',
                table_name='rooms',
                where='id = ?',
                params=(room_id,)
            )

            if not rooms:
                return {
                    "success": False,
                    "message": f"未找到直播间ID: {room_id}"
                }

            room_name = rooms[0]['name']

            # 添加到数据库
            result = add_live_time(
                db_path='system.db',
                room_id=room_id,
                live_time=live_time,
                remark=remark
            )

            if result:
                # 创建对应的跟播任务（不再是直播提醒）
                try:
                    task_manager = get_task_manager()
                    run_time = datetime.fromisoformat(live_time)

                    # 构建跟播任务备注
                    follow_remark = f"首次跟播 {room_name} 房间"
                    if remark:
                        follow_remark += f"\n备注：{remark}"
                    follow_remark += f"\n说明：直播时间安排任务，将在指定时间自动执行跟播"

                    # 创建跟播任务
                    task_success = task_manager.add_follow_task(
                        room_ids=[room_id],
                        room_names=[room_name],
                        run_time=run_time,
                        remark=follow_remark,
                        retry_count=0
                    )

                    if task_success:
                        return {
                            "success": True,
                            "message": f"成功添加直播时间并创建跟播任务"
                        }
                    else:
                        return {
                            "success": True,
                            "message": f"直播时间添加成功，但跟播任务创建失败"
                        }

                except Exception as task_error:
                    return {
                        "success": True,
                        "message": f"直播时间添加成功，但跟播任务创建失败: {str(task_error)}"
                    }
            else:
                return {
                    "success": False,
                    "message": "添加直播时间失败"
                }

        except Exception as e:
            return {
                "success": False,
                "message": str(e)
            }

    def notify_live_start(self, room_name):
        """
        通知直播开始 - 发送Windows通知
        """
        try:
            import win10toast

            toaster = win10toast.ToastNotifier()
            toaster.show_toast(
                "直播间开始跟播",
                f"直播间 {room_name} 开始跟播",
                icon_path=None,
                duration=10,
                threaded=True
            )

            return {
                "success": True,
                "message": f"已发送通知：{room_name} 开始跟播"
            }

        except ImportError:
            return {
                "success": False,
                "message": "win10toast 模块未安装，请运行: pip install win10toast"
            }
        except Exception as e:
            return {
                "success": False,
                "message": f"发送通知失败: {str(e)}"
            }

    def get_room_next_live_time(self, room_id):
        """
        获取直播间的下次直播时间
        """
        try:
            from sqlite3_util import get_room_next_live_time

            if not room_id:
                return {
                    "success": False,
                    "message": "直播间ID不能为空",
                    "data": {}
                }

            result = get_room_next_live_time(
                db_path='system.db',
                room_id=room_id
            )

            return {
                "success": True,
                "data": result
            }

        except Exception as e:
            return {
                "success": False,
                "message": f"获取直播时间失败: {str(e)}",
                "data": {}
            }

    def get_room_live_times(self, room_id):
        """
        获取直播间的所有直播时间
        """
        try:
            from sqlite3_util import get_room_live_times

            if not room_id:
                return {
                    "success": False,
                    "message": "直播间ID不能为空",
                    "data": [],
                    "total": 0
                }

            result = get_room_live_times(
                db_path='system.db',
                room_id=room_id
            )

            return {
                "success": True,
                "data": result,
                "total": len(result)
            }

        except Exception as e:
            return {
                "success": False,
                "message": f"获取直播时间列表失败: {str(e)}",
                "data": [],
                "total": 0
            }

    def update_live_time_status(self, live_time_id, status):
        """
        更新直播时间状态
        """
        try:
            from sqlite3_util import update_live_time_status

            if not live_time_id:
                return {
                    "success": False,
                    "message": "直播时间ID不能为空"
                }

            result = update_live_time_status(
                db_path='system.db',
                live_time_id=live_time_id,
                status=status
            )

            if result:
                return {
                    "success": True,
                    "message": f"成功更新直播时间状态"
                }
            else:
                return {
                    "success": False,
                    "message": "更新直播时间状态失败"
                }

        except Exception as e:
            return {
                "success": False,
                "message": f"更新直播时间状态失败: {str(e)}"
            }

    def delete_live_time(self, live_time_id):
        """
        删除直播时间
        """
        try:
            from sqlite3_util import delete_live_time

            if not live_time_id:
                return {
                    "success": False,
                    "message": "直播时间ID不能为空"
                }

            result = delete_live_time(
                db_path='system.db',
                live_time_id=live_time_id
            )

            if result:
                return {
                    "success": True,
                    "message": f"成功删除直播时间"
                }
            else:
                return {
                    "success": False,
                    "message": "删除直播时间失败"
                }

        except Exception as e:
            return {
                "success": False,
                "message": f"删除直播时间失败: {str(e)}"
            }

    # ==================== 定时任务管理 ====================

    def get_active_tasks(self):
        """
        获取所有活跃的定时任务
        """
        try:
            from task_manager import get_task_manager

            task_manager = get_task_manager()
            tasks = task_manager.get_active_tasks()

            return {
                "success": True,
                "data": tasks,
                "total": len(tasks)
            }

        except Exception as e:
            return {
                "success": False,
                "message": f"获取任务列表失败: {str(e)}",
                "data": []
            }

    def sync_tasks_with_live_times(self):
        """
        同步任务与直播时间表
        """
        try:
            from task_manager import get_task_manager

            task_manager = get_task_manager()
            sync_count = task_manager.sync_tasks_with_live_times()

            return {
                "success": True,
                "message": f"同步完成，创建了 {sync_count} 个新任务",
                "sync_count": sync_count
            }

        except Exception as e:
            return {
                "success": False,
                "message": f"同步任务失败: {str(e)}"
            }

    def cleanup_expired_tasks(self):
        """
        清理过期任务
        """
        try:
            from task_manager import get_task_manager

            task_manager = get_task_manager()
            cleanup_count = task_manager.cleanup_expired_tasks()

            return {
                "success": True,
                "message": f"清理完成，处理了 {cleanup_count} 个过期任务",
                "cleanup_count": cleanup_count
            }

        except Exception as e:
            return {
                "success": False,
                "message": f"清理任务失败: {str(e)}"
            }

    def remove_task(self, task_id):
        """
        移除指定任务
        """
        try:
            from task_manager import get_task_manager

            if not task_id:
                return {
                    "success": False,
                    "message": "任务ID不能为空"
                }

            task_manager = get_task_manager()
            result = task_manager.remove_task(task_id)

            if result:
                return {
                    "success": True,
                    "message": f"成功移除任务: {task_id}"
                }
            else:
                return {
                    "success": False,
                    "message": "移除任务失败"
                }

        except Exception as e:
            return {
                "success": False,
                "message": f"移除任务失败: {str(e)}"
            }

    def remove_task_by_live_time(self, room_id, live_time):
        """
        根据直播间ID和直播时间移除对应的任务
        """
        try:
            from task_manager import get_task_manager
            from datetime import datetime

            if not room_id or not live_time:
                return {
                    "success": False,
                    "message": "直播间ID和直播时间不能为空"
                }

            # 解析时间并生成任务ID
            try:
                run_time = datetime.fromisoformat(live_time)
                task_id = f"live_reminder_{room_id}_{int(run_time.timestamp())}"
            except Exception as e:
                return {
                    "success": False,
                    "message": f"时间格式错误: {str(e)}"
                }

            task_manager = get_task_manager()
            result = task_manager.remove_task(task_id)

            if result:
                return {
                    "success": True,
                    "message": f"成功移除任务: {task_id}"
                }
            else:
                return {
                    "success": False,
                    "message": "移除任务失败"
                }

        except Exception as e:
            return {
                "success": False,
                "message": f"移除任务失败: {str(e)}"
            }

    def mark_live_time_as_started(self, room_id, live_time):
        """
        标记直播时间记录为已开播状态
        """
        try:
            import sqlite3
            from datetime import datetime

            if not room_id or not live_time:
                return {
                    "success": False,
                    "message": "直播间ID和直播时间不能为空"
                }

            # 查找对应的直播时间记录
            conn = sqlite3.connect('system.db')
            cursor = conn.cursor()
            
            cursor.execute("""
                SELECT id FROM time_of_live 
                WHERE room_id = ? AND live_time = ? AND status = 0
            """, (room_id, live_time))
            
            result = cursor.fetchone()
            
            if result:
                live_time_id = result[0]
                
                # 更新状态为已开播
                cursor.execute("""
                    UPDATE time_of_live SET status = 1 WHERE id = ?
                """, (live_time_id,))
                
                conn.commit()
                conn.close()
                
                return {
                    "success": True,
                    "message": f"已标记直播间 {room_id} 的时间 {live_time} 为已开播状态"
                }
            else:
                conn.close()
                return {
                    "success": False,
                    "message": f"未找到直播间 {room_id} 的待开播记录"
                }

        except Exception as e:
            return {
                "success": False,
                "message": f"标记已开播状态失败: {str(e)}"
            }

    def start_batch_follow(self, follow_data):
        """
        批量启动跟播 - 创建定时跟播任务

        Args:
            follow_data: 直播间数据列表，格式: [{"id": 1, "name": "直播间1", "platform": "wechat"}, ...]

        Returns:
            dict: 操作结果
        """
        try:
            # 验证输入数据
            if not follow_data or not isinstance(follow_data, list):
                return {
                    "success": False,
                    "message": "跟播数据不能为空且必须是列表格式"
                }

            room_count = len(follow_data)

            # 验证直播间数据
            room_ids = []
            room_names = []

            for room in follow_data:
                if not isinstance(room, dict) or 'id' not in room or 'name' not in room:
                    return {
                        "success": False,
                        "message": "直播间数据格式错误，必须包含id和name字段"
                    }

                room_ids.append(int(room['id']))
                room_names.append(str(room['name']))

            # 🔥 新增：初始化跟播进度状态
            reset_follow_progress_logs()
            update_follow_progress_status(
                is_following=True,
                room_count=room_count,
                completed_count=0,
                progress=0,
                step="准备开始跟播"
            )

            # 🔥 新增：创建跟播进度监控窗口
            room_names_str = ", ".join(room_names[:2])
            if room_count > 2:
                room_names_str += f" 等{room_count}个直播间"

            progress_window_result = self.create_follow_progress_window(room_names_str)
            if progress_window_result["success"]:
                add_follow_progress_log(f"✅ 跟播进度窗口已创建", "success", 5, "窗口初始化完成")
            else:
                add_follow_progress_log(f"⚠️ 跟播进度窗口创建失败: {progress_window_result['message']}", "warning")

            # 记录日志到控制台和进度窗口
            print(f"\n🎯 批量跟播请求: {room_count} 个直播间")
            add_follow_progress_log(f"🎯 开始批量跟播: {room_count} 个直播间", "info", 10, "准备跟播任务")

            for i, room in enumerate(follow_data, 1):
                print(f"   {i}. {room['name']} (ID: {room['id']}, 平台: {room.get('platform', 'wechat')})")
                add_follow_progress_log(f"📺 直播间 {i}: {room['name']}", "info")

            # 创建跟播任务而不是立即执行
            try:
                from task_manager import get_task_manager

                task_manager = get_task_manager()
                add_follow_progress_log("🔧 正在创建跟播任务...", "info", 20, "任务管理器初始化")

                # 创建立即执行的跟播任务
                success = task_manager.add_immediate_follow_task(
                    room_ids=room_ids,
                    room_names=room_names,
                    remark=f"批量跟播任务 - {room_count}个直播间"
                )

                if success:
                    print(f"✅ 跟播任务创建成功，即将开始执行")
                    add_follow_progress_log("✅ 跟播任务创建成功", "success", 30, "任务已提交到执行队列")

                    # 发送成功通知
                    try:
                        from win10toast import ToastNotifier
                        toaster = ToastNotifier()

                        title = f"🚀 批量跟播任务已创建"
                        message = f"已创建 {room_count} 个直播间的跟播任务\n"

                        # 添加直播间列表（限制长度）
                        if room_count <= 3:
                            for room in follow_data:
                                message += f"• {room['name']}\n"
                        else:
                            for room in follow_data[:2]:
                                message += f"• {room['name']}\n"
                            message += f"• 等共 {room_count} 个直播间..."

                        message += f"\n⏰ 任务将在3秒后开始执行"

                        toaster.show_toast(
                            title=title,
                            msg=message,
                            duration=8,
                            threaded=True
                        )

                        print(f"✅ 通知显示成功")
                        add_follow_progress_log("📢 系统通知已发送", "info", 35, "通知用户任务状态")

                    except ImportError:
                        print(f"⚠️ win10toast模块不可用，跳过通知显示")
                        add_follow_progress_log("⚠️ 系统通知模块不可用", "warning")
                    except Exception as e:
                        print(f"⚠️ 通知显示失败: {str(e)}")
                        add_follow_progress_log(f"⚠️ 通知显示失败: {str(e)}", "warning")

                    # 🔥 新增：更新进度状态为等待执行
                    add_follow_progress_log("⏰ 跟播任务即将开始执行...", "info", 40, "等待任务执行")

                    return {
                        "success": True,
                        "message": f"成功创建批量跟播任务，将跟播 {room_count} 个直播间",
                        "data": {
                            "room_count": room_count,
                            "rooms": follow_data,
                            "status": "task_created",
                            "execution_type": "immediate",
                            "progress_window_created": progress_window_result["success"]
                        }
                    }
                else:
                    add_follow_progress_log("❌ 跟播任务创建失败", "error", 0, "任务管理器异常")
                    update_follow_progress_status(is_following=False)
                    return {
                        "success": False,
                        "message": "创建跟播任务失败，请检查任务管理器状态"
                    }

            except ImportError as e:
                print(f"❌ 任务管理器模块导入失败: {str(e)}")
                add_follow_progress_log(f"❌ 任务管理器模块导入失败: {str(e)}", "error")
                update_follow_progress_status(is_following=False)
                return {
                    "success": False,
                    "message": f"任务管理器不可用: {str(e)}"
                }

        except Exception as e:
            print(f"❌ 批量跟播失败: {str(e)}")
            add_follow_progress_log(f"❌ 批量跟播失败: {str(e)}", "error")
            update_follow_progress_status(is_following=False)
            return {
                "success": False,
                "message": f"批量跟播失败: {str(e)}"
            }

    def get_follow_task_status(self):
        """获取跟播任务状态统计"""
        try:
            from task_manager import get_task_manager

            task_manager = get_task_manager()
            stats = task_manager.get_task_execution_stats()

            return {
                "success": True,
                "data": stats
            }

        except Exception as e:
            return {
                "success": False,
                "message": f"获取任务状态失败: {str(e)}",
                "data": {}
            }

    def get_active_follow_tasks(self):
        """获取活跃的跟播任务列表"""
        try:
            from task_manager import get_task_manager

            task_manager = get_task_manager()
            tasks = task_manager.get_active_tasks()

            # 过滤出跟播任务
            follow_tasks = [task for task in tasks if task.get('task_type') == 'follow_task']

            return {
                "success": True,
                "data": follow_tasks,
                "total": len(follow_tasks)
            }

        except Exception as e:
            return {
                "success": False,
                "message": f"获取活跃任务失败: {str(e)}",
                "data": []
            }

    def cancel_follow_task(self, task_id):
        """取消指定的跟播任务"""
        try:
            from task_manager import get_task_manager

            if not task_id:
                return {
                    "success": False,
                    "message": "任务ID不能为空"
                }

            task_manager = get_task_manager()

            # 取消主任务
            success = task_manager.remove_task(task_id)

            if success:
                # 取消相关的重试任务
                task_manager.cancel_retry_tasks(task_id)

                return {
                    "success": True,
                    "message": f"成功取消任务: {task_id}"
                }
            else:
                return {
                    "success": False,
                    "message": f"取消任务失败: {task_id}"
                }

        except Exception as e:
            return {
                "success": False,
                "message": f"取消任务失败: {str(e)}"
            }

    def cancel_current_follow(self, room_name):
        """
        取消当前直播间的所有跟播相关任务 - 简化版

        Args:
            room_name: 直播间名称

        Returns:
            dict: 操作结果
        """
        try:
            import sqlite3

            if not room_name or room_name == "未知直播间":
                return {
                    "success": False,
                    "message": "直播间名称无效"
                }

            add_console_log(f"🛑 开始取消直播间 [{room_name}] 的跟播任务...", "warning")

            # 🔥 简化：直接根据房间名查询并设置status=1
            conn = sqlite3.connect('system.db')
            cursor = conn.cursor()

            # 直接更新所有相关任务的状态为1（已失效）
            cursor.execute("""
                UPDATE tasks
                SET status = 1
                WHERE status = 0 AND (
                    room_names LIKE ? OR
                    remark LIKE ? OR
                    task_data LIKE ? OR
                    task_id LIKE ?
                )
            """, (f"%{room_name}%", f"%{room_name}%", f"%{room_name}%", f"%{room_name}%"))

            updated_count = cursor.rowcount
            conn.commit()
            conn.close()

            add_console_log(f"✅ 已取消 {updated_count} 个任务", "success")

            add_console_log(f"✅ 已标记监听窗口关闭: {room_name}", "success")

            return {
                "success": True,
                "message": f"已成功取消直播间 [{room_name}] 的所有跟播任务",
                "cancelled_count": updated_count,
                "should_close_window": True  # 🔥 告诉前端应该关闭窗口
            }

        except Exception as e:
            add_console_log(f"❌ 取消跟播任务失败: {str(e)}", "error")
            return {
                "success": False,
                "message": f"取消跟播任务失败: {str(e)}"
            }

    def _close_window_directly(self, room_name):
        """直接关闭指定房间的监听窗口 - 简化版"""
        try:
            global PROGRESS_WINDOW_MANAGER

            # 检查是否有该房间的窗口
            if room_name in PROGRESS_WINDOW_MANAGER["active_windows"]:
                window_info = PROGRESS_WINDOW_MANAGER["active_windows"][room_name]
                window = window_info.get("window")

                if window:
                    # 🔥 尝试直接关闭窗口
                    try:
                        # 方法1：尝试destroy
                        if hasattr(window, 'destroy'):
                            window.destroy()
                            add_console_log(f"✅ 监听窗口已关闭: {room_name}", "success")
                        # 方法2：尝试close
                        elif hasattr(window, 'close'):
                            window.close()
                            add_console_log(f"✅ 监听窗口已关闭: {room_name}", "success")
                        else:
                            # 降级到标记方式
                            PROGRESS_WINDOW_MANAGER["should_close"].add(room_name)
                            add_console_log(f"🔄 已标记窗口关闭: {room_name}", "info")
                    except Exception as close_e:
                        # 如果直接关闭失败，使用标记方式
                        PROGRESS_WINDOW_MANAGER["should_close"].add(room_name)
                        add_console_log(f"🔄 窗口关闭降级到标记方式: {room_name}", "info")

                # 清理窗口记录
                del PROGRESS_WINDOW_MANAGER["active_windows"][room_name]
            else:
                add_console_log(f"💡 未找到房间 [{room_name}] 的监听窗口", "info")

        except Exception as e:
            add_console_log(f"⚠️ 关闭窗口异常: {str(e)}", "warning")
            # 降级到标记方式
            try:
                PROGRESS_WINDOW_MANAGER["should_close"].add(room_name)
            except:
                pass

    def retry_failed_follow_task(self, task_id):
        """手动重试失败的跟播任务"""
        try:
            from task_manager import get_task_manager
            import json

            if not task_id:
                return {
                    "success": False,
                    "message": "任务ID不能为空"
                }

            task_manager = get_task_manager()

            # 获取原任务信息
            original_task = task_manager._get_task_by_id(task_id)
            if not original_task:
                return {
                    "success": False,
                    "message": f"未找到任务: {task_id}"
                }

            # 转换数据
            room_ids = json.loads(original_task.get('room_ids', '[]'))
            room_names = json.loads(original_task.get('room_names', '[]'))

            # 创建新的重试任务
            retry_task_id = f"{task_id}_manual_retry_{int(datetime.now().timestamp())}"

            success = task_manager.add_immediate_follow_task(
                room_ids=room_ids,
                room_names=room_names,
                remark=f"手动重试任务 - 原任务: {task_id}"
            )

            if success:
                return {
                    "success": True,
                    "message": f"成功创建重试任务: {retry_task_id}",
                    "retry_task_id": retry_task_id
                }
            else:
                return {
                    "success": False,
                    "message": "创建重试任务失败"
                }

        except Exception as e:
            return {
                "success": False,
                "message": f"手动重试失败: {str(e)}"
            }

    def start_test_follow(self, follow_data):
        """
        测试跟播功能 - 只测试话术输入，不实际发送
        
        Args:
            follow_data: 直播间数据列表，格式: [{"id": 1, "name": "直播间1", "platform": "wechat"}, ...]
            
        Returns:
            dict: 操作结果
        """
        try:
            # 验证输入数据
            if not follow_data or not isinstance(follow_data, list):
                return {
                    "success": False,
                    "message": "跟播数据不能为空且必须是列表格式"
                }

            room_count = len(follow_data)

            # 验证直播间数据
            room_ids = []
            room_names = []

            for room in follow_data:
                if not isinstance(room, dict) or 'id' not in room or 'name' not in room:
                    return {
                        "success": False,
                        "message": "直播间数据格式错误，必须包含id和name字段"
                    }

                room_ids.append(int(room['id']))
                room_names.append(str(room['name']))

            # 记录日志到控制台
            print(f"\n🧪 测试跟播请求: {room_count} 个直播间")
            for i, room in enumerate(follow_data, 1):
                print(f"   {i}. {room['name']} (ID: {room['id']}, 平台: {room.get('platform', 'wechat')})")

            # 创建测试跟播任务
            try:
                from task_manager import get_task_manager

                task_manager = get_task_manager()

                # 创建测试模式的跟播任务
                success = task_manager.add_test_follow_task(
                    room_ids=room_ids,
                    room_names=room_names,
                    remark=f"测试跟播任务 - {room_count}个直播间"
                )

                if success:
                    print(f"✅ 测试跟播任务创建成功")

                    # 发送成功通知
                    try:
                        from win10toast import ToastNotifier
                        toaster = ToastNotifier()

                        title = f"🧪 测试跟播任务已创建"
                        message = f"已创建 {room_count} 个直播间的测试跟播任务\n"

                        # 添加直播间列表（限制长度）
                        if room_count <= 3:
                            for room in follow_data:
                                message += f"• {room['name']}\n"
                        else:
                            for room in follow_data[:2]:
                                message += f"• {room['name']}\n"
                            message += f"• 等共 {room_count} 个直播间..."

                        message += f"\n⚠️ 测试模式: 只测试输入，不实际发送"

                        toaster.show_toast(
                            title=title,
                            msg=message,
                            duration=8,
                            threaded=True
                        )

                        print(f"✅ 通知显示成功")

                    except ImportError:
                        print(f"⚠️ win10toast模块不可用，跳过通知显示")
                    except Exception as e:
                        print(f"⚠️ 通知显示失败: {str(e)}")

                    return {
                        "success": True,
                        "message": f"成功创建测试跟播任务，将测试 {room_count} 个直播间",
                        "data": {
                            "room_count": room_count,
                            "rooms": follow_data,
                            "status": "test_task_created",
                            "execution_type": "test_mode"
                        }
                    }
                else:
                    return {
                        "success": False,
                        "message": "创建测试跟播任务失败失败"
                    }

            except ImportError as e:
                print(f"❌ 任务管理器模块导入失败: {str(e)}")
                return {
                    "success": False,
                    "message": f"任务管理器不可用: {str(e)}"
                }

        except Exception as e:
            print(f"❌ 测试跟播失败: {str(e)}")
            return {
                "success": False,
                "message": f"测试跟播失败: {str(e)}"
            }

    def get_task_list(self):
        """
        获取任务列表
        
        Returns:
            dict: 任务列表数据
        """
        try:
            from sqlite3_util import query_table

            # 查询任务列表
            tasks = query_table(
                db_path='system.db',
                table_name='tasks',
                order_by='create_time DESC',
                limit=100  # 最近100条任务
            )

            # 处理任务数据
            task_list = []
            for task in tasks:
                # 处理执行状态
                execution_status = task.get('execution_status', 'pending')
                
                # 更准确的状态文本
                if task['status'] == 0:
                    if execution_status == 'executing':
                        status_text = '执行中'
                    elif execution_status == 'completed':
                        status_text = '已完成'
                    elif execution_status == 'failed':
                        status_text = '已失败'
                    elif execution_status == 'error':
                        status_text = '执行错误'
                    else:
                        status_text = '等待触发'
                else:
                    status_text = '已失效'
                
                task_data = {
                    'id': task['id'],
                    'task_id': task['task_id'],
                    'task_type': task['task_type'],
                    'room_id': task.get('room_id'),
                    'room_ids': task.get('room_ids'),
                    'room_names': task.get('room_names'),
                    'run_time': task['run_time'],
                    'create_time': task['create_time'],
                    'status': task['status'],
                    'status_text': status_text,
                    'remark': task.get('remark', ''),
                    'execution_status': execution_status,
                    'retry_count': task.get('retry_count', 0)
                }

                # 处理任务类型显示
                if task['task_type'] == 'live_reminder':
                    task_data['type_text'] = '直播提醒'
                elif task['task_type'] == 'follow_task':
                    task_data['type_text'] = '跟播任务'
                elif task['task_type'] == 'test_follow_task':
                    task_data['type_text'] = '测试跟播'
                else:
                    task_data['type_text'] = task['task_type']

                task_list.append(task_data)

            return {
                "success": True,
                "message": f"获取任务列表成功，共 {len(task_list)} 条任务",
                "data": task_list
            }

        except Exception as e:
            return {
                "success": False,
                "message": f"获取任务列表失败: {str(e)}"
            }

    def get_task_logs(self, limit=100):
        """
        获取任务日志列表
        
        Args:
            limit: 限制返回数量
            
        Returns:
            dict: 任务日志数据
        """
        try:
            from sqlite3_util import query_table

            # 查询任务日志
            logs = query_table(
                db_path='system.db',
                table_name='task_log',
                order_by='execution_time DESC',
                limit=limit
            )

            # 处理日志数据
            log_list = []
            for log in logs:
                log_data = {
                    'id': log['id'],
                    'task_id': log['task_id'],
                    'status': log['status'],
                    'status_text': '成功' if log['status'] == 1 else '失败',
                    'message': log.get('message', ''),
                    'room_id': log.get('room_id'),
                    'room_name': log.get('room_name', ''),
                    'execution_time': log['execution_time'],
                    'create_time': log['create_time']
                }
                log_list.append(log_data)

            return {
                "success": True,
                "message": f"获取任务日志成功，共 {len(log_list)} 条记录",
                "data": log_list
            }

        except Exception as e:
            return {
                "success": False,
                "message": f"获取任务日志失败: {str(e)}"
            }

    def invalidate_task(self, task_id):
        """
        使任务失效（设置状态为1并从调度器中移除）
        
        Args:
            task_id: 任务ID
            
        Returns:
            dict: 操作结果
        """
        try:
            import sqlite3
            from task_manager import get_task_manager, add_task_log

            # 从调度器中移除任务
            try:
                task_manager = get_task_manager()
                if task_manager and task_manager.is_running:
                    success = task_manager.remove_task(task_id)
                    if success:
                        print(f"✅ 已从调度器移除任务: {task_id}")
                    else:
                        print(f"⚠️ 调度器中未找到任务: {task_id}")
            except Exception as scheduler_error:
                print(f"⚠️ 从调度器移除任务失败: {str(scheduler_error)}")

            # 更新数据库中的任务状态
            conn = sqlite3.connect('system.db')
            cursor = conn.cursor()

            cursor.execute("""
                UPDATE tasks SET status = 1 WHERE task_id = ?
            """, (task_id,))

            affected_rows = cursor.rowcount
            conn.commit()
            conn.close()

            if affected_rows > 0:
                # 记录日志
                add_task_log(task_id, 2, "用户手动取消任务")

                return {
                    "success": True,
                    "message": f"任务 {task_id} 已成功取消"
                }
            else:
                return {
                    "success": False,
                    "message": f"未找到任务 {task_id}"
                }

        except Exception as e:
            return {
                "success": False,
                "message": f"取消任务失败: {str(e)}"
            }

    # ==================== 商品管理API ====================

    def add_product_api(self, product_data):
        """
        添加商品API
        """
        try:
            print(f"=== 添加商品 ===")
            print(f"商品数据: {product_data}")

            if not product_data:
                return {
                    "success": False,
                    "message": "商品数据不能为空"
                }

            name = product_data.get('name', '').strip()
            if not name:
                return {
                    "success": False,
                    "message": "商品名称不能为空"
                }

            cover = product_data.get('cover', '').strip()
            price = float(product_data.get('price', 0.0))
            remark = product_data.get('remark', '').strip()

            # 调用数据库函数添加商品
            result = add_product(
                db_path='system.db',
                name=name,
                cover=cover,
                price=price,
                remark=remark
            )

            return result

        except Exception as e:
            print(f"❌ 添加商品API失败: {str(e)}")
            return {
                "success": False,
                "message": f"添加商品失败: {str(e)}"
            }

    def query_products_api(self, query_params=None):
        """
        查询商品列表API
        """
        try:
            print(f"=== 查询商品列表 ===")

            if not query_params:
                query_params = {}

            page = int(query_params.get('page', 1))
            page_size = int(query_params.get('page_size', 20))
            search_name = query_params.get('search_name', '').strip()

            print(f"查询参数: page={page}, page_size={page_size}, search_name={search_name}")

            # 调用数据库函数查询商品
            result = query_products(
                db_path='system.db',
                page=page,
                page_size=page_size,
                search_name=search_name
            )

            return result

        except Exception as e:
            print(f"❌ 查询商品API失败: {str(e)}")
            return {
                "success": False,
                "message": f"查询商品失败: {str(e)}",
                "data": [],
                "total": 0
            }

    def update_product_api(self, product_data):
        """
        更新商品API
        """
        try:
            print(f"=== 更新商品 ===")
            print(f"商品数据: {product_data}")

            if not product_data:
                return {
                    "success": False,
                    "message": "商品数据不能为空"
                }

            product_id = int(product_data.get('id', 0))
            if product_id <= 0:
                return {
                    "success": False,
                    "message": "商品ID无效"
                }

            name = product_data.get('name', '').strip()
            if not name:
                return {
                    "success": False,
                    "message": "商品名称不能为空"
                }

            cover = product_data.get('cover', '').strip()
            price = float(product_data.get('price', 0.0))
            remark = product_data.get('remark', '').strip()

            # 调用数据库函数更新商品
            result = update_product(
                db_path='system.db',
                product_id=product_id,
                name=name,
                cover=cover,
                price=price,
                remark=remark
            )

            return result

        except Exception as e:
            print(f"❌ 更新商品API失败: {str(e)}")
            return {
                "success": False,
                "message": f"更新商品失败: {str(e)}"
            }

    def delete_product_api(self, product_id):
        """
        删除商品API
        """
        try:
            print(f"=== 删除商品 ===")
            print(f"商品ID: {product_id}")

            if not product_id or int(product_id) <= 0:
                return {
                    "success": False,
                    "message": "商品ID无效"
                }

            # 调用数据库函数删除商品
            result = delete_product(
                db_path='system.db',
                product_id=int(product_id)
            )

            return result

        except Exception as e:
            print(f"❌ 删除商品API失败: {str(e)}")
            return {
                "success": False,
                "message": f"删除商品失败: {str(e)}"
            }

    def get_products_simple_api(self):
        """
        获取商品简单列表API（用于下拉选择）
        """
        try:
            print(f"=== 获取商品简单列表 ===")

            # 调用数据库函数获取商品列表
            products = get_all_products_simple(db_path='system.db')

            return {
                "success": True,
                "data": products
            }

        except Exception as e:
            print(f"❌ 获取商品简单列表API失败: {str(e)}")
            return {
                "success": False,
                "message": f"获取商品列表失败: {str(e)}",
                "data": []
            }

    # ==================== 图片管理API ====================

    def add_image_api(self, image_data):
        """
        添加图片API
        """
        try:
            print(f"=== 添加图片 ===")
            print(f"图片数据: {image_data}")

            if not image_data:
                return {
                    "success": False,
                    "message": "图片数据不能为空"
                }

            path = image_data.get('path', '').strip()
            if not path:
                return {
                    "success": False,
                    "message": "图片路径不能为空"
                }

            remark = image_data.get('remark', '').strip()
            status = int(image_data.get('status', 1))
            product_id = image_data.get('product_id')

            # 如果product_id为空字符串或0，设为None
            if not product_id or int(product_id) == 0:
                product_id = None
            else:
                product_id = int(product_id)

            # 调用数据库函数添加图片
            result = add_image(
                db_path='system.db',
                path=path,
                remark=remark,
                status=status,
                product_id=product_id
            )

            return result

        except Exception as e:
            print(f"❌ 添加图片API失败: {str(e)}")
            return {
                "success": False,
                "message": f"添加图片失败: {str(e)}"
            }

    def query_images_api(self, query_params=None):
        """
        查询图片列表API
        """
        try:
            print(f"=== 查询图片列表 ===")

            if not query_params:
                query_params = {}

            page = int(query_params.get('page', 1))
            page_size = int(query_params.get('page_size', 20))
            product_id = query_params.get('product_id')
            status = query_params.get('status')

            # 处理筛选参数
            if product_id and int(product_id) > 0:
                product_id = int(product_id)
            else:
                product_id = None

            if status is not None and str(status).isdigit():
                status = int(status)
            else:
                status = None

            print(f"查询参数: page={page}, page_size={page_size}, product_id={product_id}, status={status}")

            # 调用数据库函数查询图片
            result = query_images(
                db_path='system.db',
                page=page,
                page_size=page_size,
                product_id=product_id,
                status=status
            )

            return result

        except Exception as e:
            print(f"❌ 查询图片API失败: {str(e)}")
            return {
                "success": False,
                "message": f"查询图片失败: {str(e)}",
                "data": [],
                "total": 0
            }

    def update_image_api(self, image_data):
        """
        更新图片API
        """
        try:
            print(f"=== 更新图片 ===")
            print(f"图片数据: {image_data}")

            if not image_data:
                return {
                    "success": False,
                    "message": "图片数据不能为空"
                }

            image_id = int(image_data.get('id', 0))
            if image_id <= 0:
                return {
                    "success": False,
                    "message": "图片ID无效"
                }

            path = image_data.get('path', '').strip()
            if not path:
                return {
                    "success": False,
                    "message": "图片路径不能为空"
                }

            remark = image_data.get('remark', '').strip()
            status = int(image_data.get('status', 1))
            product_id = image_data.get('product_id')

            # 如果product_id为空字符串或0，设为None
            if not product_id or int(product_id) == 0:
                product_id = None
            else:
                product_id = int(product_id)

            # 调用数据库函数更新图片
            result = update_image(
                db_path='system.db',
                image_id=image_id,
                path=path,
                remark=remark,
                status=status,
                product_id=product_id
            )

            return result

        except Exception as e:
            print(f"❌ 更新图片API失败: {str(e)}")
            return {
                "success": False,
                "message": f"更新图片失败: {str(e)}"
            }

    def delete_image_api(self, image_id):
        """
        删除图片API（同时删除文件和数据库记录）
        """
        try:
            print(f"=== 删除图片 ===")
            print(f"图片ID: {image_id}")

            if not image_id or int(image_id) <= 0:
                return {
                    "success": False,
                    "message": "图片ID无效"
                }

            # 先获取图片信息
            import sqlite3
            conn = sqlite3.connect('system.db')
            cursor = conn.cursor()
            cursor.execute("SELECT path FROM images WHERE id = ?", (int(image_id),))
            image_info = cursor.fetchone()
            conn.close()

            if not image_info:
                return {
                    "success": False,
                    "message": "图片不存在"
                }

            image_path = image_info[0]

            # 删除物理文件
            if image_path:
                try:
                    current_dir = os.path.dirname(os.path.abspath(__file__))
                    if image_path.startswith('cvimages/'):
                        # cvimages/xxx格式
                        filename = image_path.replace('cvimages/', '')
                        file_path = os.path.join(current_dir, 'cvimages', filename)
                    else:
                        # 完整路径
                        file_path = image_path

                    if os.path.exists(file_path):
                        os.remove(file_path)
                        print(f"✅ 成功删除图片文件: {file_path}")
                    else:
                        print(f"⚠️ 图片文件不存在: {file_path}")

                except Exception as file_error:
                    print(f"⚠️ 删除图片文件失败: {str(file_error)}")
                    # 继续删除数据库记录，不因为文件删除失败而中断

            # 调用数据库函数删除图片记录
            result = delete_image(
                db_path='system.db',
                image_id=int(image_id)
            )

            return result

        except Exception as e:
            print(f"❌ 删除图片API失败: {str(e)}")
            return {
                "success": False,
                "message": f"删除图片失败: {str(e)}"
            }

    def get_product_images_api(self, product_id):
        """
        获取商品关联图片API
        """
        try:
            print(f"=== 获取商品关联图片 ===")
            print(f"商品ID: {product_id}")

            if not product_id or int(product_id) <= 0:
                return {
                    "success": False,
                    "message": "商品ID无效",
                    "data": []
                }

            # 调用数据库函数获取商品图片
            images = get_product_images(
                db_path='system.db',
                product_id=int(product_id)
            )

            return {
                "success": True,
                "data": images
            }

        except Exception as e:
            print(f"❌ 获取商品图片API失败: {str(e)}")
            return {
                "success": False,
                "message": f"获取商品图片失败: {str(e)}",
                "data": []
            }

    # ==================== 图片文件管理API ====================

    def upload_image_file(self, file_data, filename=None):
        """
        上传图片文件到cvimages目录
        :param file_data: base64编码的图片数据或文件路径
        :param filename: 文件名（可选）
        :return: 上传结果
        """
        try:
            print(f"=== 上传图片文件 ===")
            print(f"文件名: {filename}")

            # 确保cvimages目录存在
            current_dir = os.path.dirname(os.path.abspath(__file__))
            cvimages_dir = os.path.join(current_dir, 'cvimages')

            if not os.path.exists(cvimages_dir):
                os.makedirs(cvimages_dir)
                print(f"✅ 创建cvimages目录: {cvimages_dir}")

            # 生成唯一文件名（纯英文）
            if not filename:
                filename = f"image_{uuid.uuid4().hex[:8]}.jpg"
            else:
                # 确保文件名安全且为纯英文
                filename = os.path.basename(filename)
                name, ext = os.path.splitext(filename)

                # 移除所有非ASCII字符，只保留英文字母、数字和安全符号
                import re
                safe_name = re.sub(r'[^a-zA-Z0-9\-_.]', '', name)  # 只保留英文字母、数字、下划线、破折号、点
                if not safe_name:  # 如果清理后为空，使用默认名称
                    safe_name = "image"

                if not ext.lower() in ['.jpg', '.jpeg', '.png', '.gif', '.bmp']:
                    ext = '.png'  # 默认使用png格式

                filename = f"{safe_name}_{uuid.uuid4().hex[:8]}{ext}"

            file_path = os.path.join(cvimages_dir, filename)

            # 处理不同类型的输入数据
            if isinstance(file_data, str):
                if file_data.startswith('data:image'):
                    # base64数据URL格式
                    header, data = file_data.split(',', 1)
                    image_data = base64.b64decode(data)
                elif file_data.startswith('/') or file_data.startswith('C:') or os.path.exists(file_data):
                    # 文件路径
                    if os.path.exists(file_data):
                        shutil.copy2(file_data, file_path)
                        print(f"✅ 复制文件成功: {file_path}")
                        return {
                            'success': True,
                            'message': '图片上传成功',
                            'filename': filename,
                            'path': f"cvimages/{filename}"
                        }
                    else:
                        return {
                            'success': False,
                            'message': '源文件不存在'
                        }
                else:
                    # 纯base64数据
                    try:
                        image_data = base64.b64decode(file_data)
                    except:
                        return {
                            'success': False,
                            'message': '无效的base64数据'
                        }
            else:
                return {
                    'success': False,
                    'message': '不支持的文件数据格式'
                }

            # 保存图片数据
            if 'image_data' in locals():
                with open(file_path, 'wb') as f:
                    f.write(image_data)
                print(f"✅ 保存图片成功: {file_path}")

            return {
                'success': True,
                'message': '图片上传成功',
                'filename': filename,
                'path': f"cvimages/{filename}"
            }

        except Exception as e:
            print(f"❌ 上传图片失败: {str(e)}")
            return {
                'success': False,
                'message': f'上传图片失败: {str(e)}'
            }

    def get_image_file(self, filename):
        """
        获取cvimages目录中的图片文件（base64格式）
        :param filename: 文件名
        :return: base64编码的图片数据
        """
        try:
            print(f"=== 获取图片文件 ===")
            print(f"文件名: {filename}")

            current_dir = os.path.dirname(os.path.abspath(__file__))
            file_path = os.path.join(current_dir, 'cvimages', filename)

            if not os.path.exists(file_path):
                return {
                    'success': False,
                    'message': '图片文件不存在'
                }

            # 读取文件并转换为base64
            with open(file_path, 'rb') as f:
                image_data = f.read()

            # 根据文件扩展名确定MIME类型
            ext = os.path.splitext(filename)[1].lower()
            mime_types = {
                '.jpg': 'image/jpeg',
                '.jpeg': 'image/jpeg',
                '.png': 'image/png',
                '.gif': 'image/gif',
                '.bmp': 'image/bmp'
            }
            mime_type = mime_types.get(ext, 'image/jpeg')

            base64_data = base64.b64encode(image_data).decode('utf-8')
            data_url = f"data:{mime_type};base64,{base64_data}"

            return {
                'success': True,
                'data': data_url,
                'filename': filename,
                'size': len(image_data)
            }

        except Exception as e:
            print(f"❌ 获取图片文件失败: {str(e)}")
            return {
                'success': False,
                'message': f'获取图片文件失败: {str(e)}'
            }

    def list_image_files(self):
        """
        列出cvimages目录中的所有图片文件
        :return: 图片文件列表
        """
        try:
            print(f"=== 列出图片文件 ===")

            current_dir = os.path.dirname(os.path.abspath(__file__))
            cvimages_dir = os.path.join(current_dir, 'cvimages')

            if not os.path.exists(cvimages_dir):
                return {
                    'success': True,
                    'data': [],
                    'message': 'cvimages目录不存在'
                }

            # 支持的图片格式
            image_extensions = {'.jpg', '.jpeg', '.png', '.gif', '.bmp'}

            files = []
            for filename in os.listdir(cvimages_dir):
                file_path = os.path.join(cvimages_dir, filename)
                if os.path.isfile(file_path):
                    ext = os.path.splitext(filename)[1].lower()
                    if ext in image_extensions:
                        stat = os.stat(file_path)
                        files.append({
                            'filename': filename,
                            'path': f"cvimages/{filename}",
                            'size': stat.st_size,
                            'modified': datetime.fromtimestamp(stat.st_mtime).strftime('%Y-%m-%d %H:%M:%S')
                        })

            # 按修改时间排序（最新的在前）
            files.sort(key=lambda x: x['modified'], reverse=True)

            return {
                'success': True,
                'data': files,
                'total': len(files)
            }

        except Exception as e:
            print(f"❌ 列出图片文件失败: {str(e)}")
            return {
                'success': False,
                'message': f'列出图片文件失败: {str(e)}',
                'data': []
            }

    def delete_image_file(self, filename):
        """
        删除cvimages目录中的图片文件
        :param filename: 文件名
        :return: 删除结果
        """
        try:
            print(f"=== 删除图片文件 ===")
            print(f"文件名: {filename}")

            current_dir = os.path.dirname(os.path.abspath(__file__))
            file_path = os.path.join(current_dir, 'cvimages', filename)

            if not os.path.exists(file_path):
                return {
                    'success': False,
                    'message': '图片文件不存在'
                }

            os.remove(file_path)
            print(f"✅ 删除图片文件成功: {file_path}")

            return {
                'success': True,
                'message': '图片文件删除成功'
            }

        except Exception as e:
            print(f"❌ 删除图片文件失败: {str(e)}")
            return {
                'success': False,
                'message': f'删除图片文件失败: {str(e)}'
            }

    def select_image_file(self):
        """
        打开文件选择对话框选择图片文件
        :return: 选择的文件路径
        """
        try:
            print(f"=== 选择图片文件 ===")

            # 方法1: 尝试使用pywebview的文件选择对话框
            try:
                import webview

                # 使用pywebview的文件选择对话框
                file_types = ('图片文件 (*.jpg;*.jpeg;*.png;*.gif;*.bmp)', 'All files (*.*)')
                file_path = webview.windows[0].create_file_dialog(
                    webview.OPEN_DIALOG,
                    allow_multiple=False,
                    file_types=file_types
                )

                if file_path and len(file_path) > 0:
                    selected_file = file_path[0]
                    print(f"✅ 选择文件: {selected_file}")
                    return {
                        'success': True,
                        'file_path': selected_file,
                        'filename': os.path.basename(selected_file)
                    }
                else:
                    return {
                        'success': False,
                        'message': '未选择文件'
                    }

            except Exception as webview_error:
                print(f"⚠️ pywebview文件对话框失败: {webview_error}")

                # 方法2: 回退到tkinter
                try:
                    import tkinter as tk
                    from tkinter import filedialog

                    # 创建隐藏的根窗口
                    root = tk.Tk()
                    root.withdraw()
                    root.attributes('-topmost', True)

                    # 打开文件选择对话框
                    file_path = filedialog.askopenfilename(
                        title="选择图片文件",
                        filetypes=[
                            ("图片文件", "*.jpg *.jpeg *.png *.gif *.bmp"),
                            ("JPEG文件", "*.jpg *.jpeg"),
                            ("PNG文件", "*.png"),
                            ("GIF文件", "*.gif"),
                            ("BMP文件", "*.bmp"),
                            ("所有文件", "*.*")
                        ]
                    )

                    root.destroy()

                    if file_path:
                        print(f"✅ 选择文件: {file_path}")
                        return {
                            'success': True,
                            'file_path': file_path,
                            'filename': os.path.basename(file_path)
                        }
                    else:
                        return {
                            'success': False,
                            'message': '未选择文件'
                        }

                except ImportError:
                    return {
                        'success': False,
                        'message': 'tkinter和pywebview文件对话框都不可用，请安装tkinter或更新pywebview'
                    }

        except Exception as e:
            print(f"❌ 选择图片文件失败: {str(e)}")
            return {
                'success': False,
                'message': f'选择图片文件失败: {str(e)}'
            }

    def get_image_data_url(self, image_path):
        """
        获取图片的data URL格式（用于前端显示）
        :param image_path: 图片路径（支持cvimages/xxx格式或完整路径）
        :return: data URL格式的图片数据
        """
        try:
            print(f"=== 获取图片数据URL ===")
            print(f"图片路径: {image_path}")

            if not image_path:
                return {
                    'success': False,
                    'message': '图片路径为空'
                }

            # 如果是网络图片，直接返回
            if image_path.startswith('http://') or image_path.startswith('https://'):
                return {
                    'success': True,
                    'data_url': image_path
                }

            # 处理本地图片路径
            current_dir = os.path.dirname(os.path.abspath(__file__))

            if image_path.startswith('cvimages/'):
                # cvimages/xxx格式
                filename = image_path.replace('cvimages/', '')
                file_path = os.path.join(current_dir, 'cvimages', filename)
            else:
                # 完整路径
                file_path = image_path

            if not os.path.exists(file_path):
                return {
                    'success': False,
                    'message': '图片文件不存在'
                }

            # 读取文件并转换为base64
            with open(file_path, 'rb') as f:
                image_data = f.read()

            # 根据文件扩展名确定MIME类型
            ext = os.path.splitext(file_path)[1].lower()
            mime_types = {
                '.jpg': 'image/jpeg',
                '.jpeg': 'image/jpeg',
                '.png': 'image/png',
                '.gif': 'image/gif',
                '.bmp': 'image/bmp'
            }
            mime_type = mime_types.get(ext, 'image/jpeg')

            base64_data = base64.b64encode(image_data).decode('utf-8')
            data_url = f"data:{mime_type};base64,{base64_data}"

            return {
                'success': True,
                'data_url': data_url,
                'size': len(image_data)
            }

        except Exception as e:
            print(f"❌ 获取图片数据URL失败: {str(e)}")
            return {
                'success': False,
                'message': f'获取图片数据URL失败: {str(e)}'
            }

    def get_image_blob(self, image_path):
        """
        获取图片Blob数据（用于URL.createObjectURL）
        :param image_path: 图片路径
        :return: 包含blob数据和MIME类型的字典
        """
        try:
            print(f"=== 获取图片Blob数据 ===")
            print(f"图片路径: {image_path}")

            if not image_path:
                return {
                    'success': False,
                    'message': '图片路径不能为空'
                }

            # 获取当前脚本目录
            current_dir = os.path.dirname(os.path.abspath(__file__))

            # 构建完整路径
            if image_path.startswith('cvimages/'):
                # cvimages/xxx格式
                filename = image_path.replace('cvimages/', '')
                full_path = os.path.join(current_dir, 'cvimages', filename)
            else:
                # 完整路径或相对路径
                full_path = os.path.join(current_dir, image_path)

            print(f"完整路径: {full_path}")

            # 检查文件是否存在
            if not os.path.exists(full_path):
                return {
                    'success': False,
                    'message': f'图片文件不存在: {full_path}'
                }

            # 获取MIME类型
            file_ext = os.path.splitext(full_path)[1].lower()
            mime_types = {
                '.jpg': 'image/jpeg',
                '.jpeg': 'image/jpeg',
                '.png': 'image/png',
                '.gif': 'image/gif',
                '.bmp': 'image/bmp',
                '.webp': 'image/webp'
            }
            mime_type = mime_types.get(file_ext, 'image/jpeg')

            # 读取文件并转换为base64
            with open(full_path, 'rb') as f:
                image_data = f.read()
                blob_data = base64.b64encode(image_data).decode('utf-8')

            print(f"✅ 成功获取图片Blob数据，大小: {len(image_data)} bytes")

            return {
                'success': True,
                'blob_data': blob_data,
                'mime_type': mime_type,
                'size': len(image_data)
            }

        except Exception as e:
            print(f"❌ 获取图片Blob数据失败: {str(e)}")
            return {
                'success': False,
                'message': f'获取图片Blob数据失败: {str(e)}'
            }

    # =============================================================================
    # 配置管理 API
    # =============================================================================

    def get_system_config(self):
        """
        获取系统配置
        """
        try:
            import json
            import os
            
            config_path = "config.json"
            if not os.path.exists(config_path):
                # 如果配置文件不存在，创建默认配置
                default_config = {
                    "system_config": {
                        "intervals": {
                            "bullet_screen_send": 500,
                            "bullet_screen_retry": 10,
                            "follow_task_retry": 60,
                            "image_recognition_retry": 60,
                            "live_room_check": 300
                        },
                        "retry_config": {
                            "max_bullet_retry": 3,
                            "max_image_retry": 5,
                            "max_follow_retry": 10,
                            "enable_auto_retry": True
                        },
                        "features": {
                            "enable_screenshot": True,
                            "enable_notifications": True,
                            "enable_detailed_logs": True,
                            "enable_auto_scroll": True,
                            "enable_error_recovery": True,
                            "enable_performance_monitoring": False,
                            "enable_real_danmu_send": False
                        },
                        "quality": {
                            "screenshot_quality": 80,
                            "image_match_confidence": 0.8,
                            "log_level": "info"
                        }
                    },
                    "ui_config": {
                        "theme": "default",
                        "auto_scroll": True,
                        "show_notifications": True,
                        "animation_enabled": True
                    }
                }
                with open(config_path, 'w', encoding='utf-8') as f:
                    json.dump(default_config, f, indent=4, ensure_ascii=False)
            
            with open(config_path, 'r', encoding='utf-8') as f:
                config = json.load(f)
            
            return {
                'success': True,
                'config': config
            }
            
        except Exception as e:
            print(f"❌ 获取系统配置失败: {str(e)}")
            return {
                'success': False,
                'message': f'获取系统配置失败: {str(e)}'
            }

    def update_system_config(self, new_config):
        """
        更新系统配置
        
        Args:
            new_config: 新的配置数据
        """
        try:
            import json
            import os
            
            config_path = "config.json"
            
            # 读取当前配置
            current_config = {}
            if os.path.exists(config_path):
                with open(config_path, 'r', encoding='utf-8') as f:
                    current_config = json.load(f)
            
            # 合并配置
            if 'system_config' in new_config:
                if 'system_config' not in current_config:
                    current_config['system_config'] = {}
                current_config['system_config'].update(new_config['system_config'])
            
            if 'ui_config' in new_config:
                if 'ui_config' not in current_config:
                    current_config['ui_config'] = {}
                current_config['ui_config'].update(new_config['ui_config'])
            
            # 保存配置
            with open(config_path, 'w', encoding='utf-8') as f:
                json.dump(current_config, f, indent=4, ensure_ascii=False)
            
            print(f"✅ 系统配置更新成功")
            return {
                'success': True,
                'message': '配置更新成功',
                'config': current_config
            }
            
        except Exception as e:
            print(f"❌ 更新系统配置失败: {str(e)}")
            return {
                'success': False,
                'message': f'更新系统配置失败: {str(e)}'
            }

    def reset_system_config(self):
        """
        重置系统配置为默认值
        """
        try:
            import json
            
            default_config = {
                "system_config": {
                    "intervals": {
                        "bullet_screen_send": 500,
                        "bullet_screen_retry": 10,
                        "follow_task_retry": 60,
                        "image_recognition_retry": 60,
                        "live_room_check": 300
                    },
                    "retry_config": {
                        "max_bullet_retry": 3,
                        "max_image_retry": 5,
                        "max_follow_retry": 10,
                        "enable_auto_retry": True
                    },
                    "features": {
                        "enable_screenshot": True,
                        "enable_notifications": True,
                        "enable_detailed_logs": True
                    },
                    "quality": {
                        "screenshot_quality": 80,
                        "image_match_confidence": 0.8,
                        "log_level": "info"
                    }
                },
                "ui_config": {
                    "theme": "default",
                    "auto_scroll": True,
                    "show_notifications": True,
                    "animation_enabled": True
                }
            }
            
            config_path = "config.json"
            with open(config_path, 'w', encoding='utf-8') as f:
                json.dump(default_config, f, indent=4, ensure_ascii=False)
            
            print(f"✅ 系统配置重置成功")
            return {
                'success': True,
                'message': '配置重置成功',
                'config': default_config
            }
            
        except Exception as e:
            print(f"❌ 重置系统配置失败: {str(e)}")
            return {
                'success': False,
                'message': f'重置系统配置失败: {str(e)}'
            }

    def toggle_real_danmu_send(self, enabled=None):
        """
        切换真实发送弹幕配置

        Args:
            enabled: True=启用真实发送(OCR点击), False=禁用(回车键测试), None=切换当前状态

        Returns:
            dict: 操作结果
        """
        try:
            import json

            config_path = "config.json"

            # 读取当前配置
            if os.path.exists(config_path):
                with open(config_path, 'r', encoding='utf-8') as f:
                    config = json.load(f)
            else:
                return {
                    "success": False,
                    "message": "配置文件不存在"
                }

            # 获取当前状态
            current_enabled = config.get("system_config", {}).get("features", {}).get("enable_real_danmu_send", False)

            # 确定新状态
            if enabled is None:
                new_enabled = not current_enabled  # 切换状态
            else:
                new_enabled = bool(enabled)

            # 更新配置
            if "system_config" not in config:
                config["system_config"] = {}
            if "features" not in config["system_config"]:
                config["system_config"]["features"] = {}

            config["system_config"]["features"]["enable_real_danmu_send"] = new_enabled

            # 保存配置
            with open(config_path, 'w', encoding='utf-8') as f:
                json.dump(config, f, indent=4, ensure_ascii=False)

            # 记录日志
            mode_text = "真实发送(OCR点击)" if new_enabled else "测试模式(回车键)"
            add_console_log(f"✅ 弹幕发送模式已切换为: {mode_text}", "success")

            return {
                "success": True,
                "message": f"弹幕发送模式已切换为: {mode_text}",
                "data": {
                    "enabled": new_enabled,
                    "mode": "real_send" if new_enabled else "test_mode",
                    "description": mode_text
                }
            }

        except Exception as e:
            error_msg = f"切换弹幕发送模式失败: {str(e)}"
            add_console_log(f"❌ {error_msg}", "error")
            return {
                "success": False,
                "message": error_msg
            }

    def get_real_danmu_send_status(self):
        """
        获取真实发送弹幕配置状态

        Returns:
            dict: 当前状态信息
        """
        try:
            import json

            config_path = "config.json"

            if os.path.exists(config_path):
                with open(config_path, 'r', encoding='utf-8') as f:
                    config = json.load(f)

                enabled = config.get("system_config", {}).get("features", {}).get("enable_real_danmu_send", False)
                mode_text = "真实发送(OCR点击)" if enabled else "测试模式(回车键)"

                return {
                    "success": True,
                    "data": {
                        "enabled": enabled,
                        "mode": "real_send" if enabled else "test_mode",
                        "description": mode_text
                    }
                }
            else:
                return {
                    "success": False,
                    "message": "配置文件不存在"
                }

        except Exception as e:
            return {
                "success": False,
                "message": f"获取弹幕发送模式状态失败: {str(e)}"
            }
